#!/usr/bin/env python3
"""Stage 6: schema card + metric plan -> a deployable dash-server workspace.

    python3 build_dashboard.py --card card.json --plan plan.json \
        --name revops --title "Revenue Operations" --out ./workspace

Emits parameterised SQL, an app.py carrying the derived metrics, a rule-based
insight panel, a guarded question panel and CSV download, plus a dash-app.json
declaring each query as a governed consumption output.

Nothing here is dataset-specific: every table, column, cast and filter comes
from the card.
"""
from __future__ import annotations
import argparse, json, os, re, sys
from string import Template


def setup_key_command() -> str:
    """How to invoke the key setup script, spelled for the current platform.

    A bare path works on macOS and Linux but not from PowerShell, and hardcoding
    "python3" can resolve to the Microsoft Store stub on Windows. `sys.executable`
    is the interpreter actually running, which is always the right one.
    """
    script = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "setup_llm_key.py")
    quote = '"' if " " in sys.executable or " " in script else ""
    return f"{quote}{sys.executable}{quote} {quote}{script}{quote}"


def entity_noun(model) -> str:
    """What the concentration insight should call the things it is counting.

    Hardcoding "accounts" puts "the ten largest of 45 accounts" on a dashboard
    whose entities are stores.
    """
    # "the ten largest of 2 week types" is not a ranking. Prefer a dimension
    # with enough members for concentration to mean anything.
    column = next(iter(model.entities or []), None)
    if column is None:
        ranked = sorted((c for c in model.dims if (c.get("distinct") or 0) >= 5),
                        key=lambda c: -(c.get("distinct") or 0))
        column = next(iter(ranked), None) or next(iter(model.dims or []), None)
    if not column:
        return "accounts"
    name = re.sub(r"^[A-Z]{1,3}_", "", str(column["name"]))
    text = re.sub(r"[_\-]+", " ", name).strip().lower()
    text = re.sub(r"\s*\b(id|key|no|number|code)\b\s*$", "", text).strip() or text
    return text if text.endswith("s") else text + "s"


def measure_kind(column: dict) -> str:
    """money | count -- decides whether a figure gets a currency symbol."""
    if column.get("_kind"):
        return column["_kind"]
    if column.get("_expr"):
        expression = column["_expr"].lower()
        if expression.strip().isdigit() or "then 1 else 0" in expression:
            return "count"
        if any(word in expression for word in
               ("price", "cost", "amount", "revenue", "sales", "charge", "profit")):
            return "money"
        return "count"
    if column["semantic_role"] == "monetary_rate":
        return "percent"          # a discount is a fraction, shown as 14.3%
    return "money" if column["semantic_role"] == "monetary_amount" else "count"


def q(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


class Model:
    """The handful of card facts the templates need, chosen once."""

    def __init__(self, card: dict, plan: dict):
        self.card = card
        self.plan = plan
        self.schema = card["schema"]
        facts = [t for t in card["tables"] if t["role"] in ("fact_like", "flat")]
        self.fact = max(facts or card["tables"], key=lambda t: t["row_count"])
        cols = self.fact["columns"]
        self.by_name = {c["name"]: c for c in cols}

        self.time = next((c for c in cols if c["semantic_role"] == "temporal_event"), None)
        self.measures = [c for c in cols if c["additivity"] in ("additive", "count_or_rate")][:4]
        self.rates = [c for c in cols if c["additivity"] == "non_additive"][:2]
        # Not truncated here: prefer_slices is applied further down, and cutting
        # to six by column order first would drop a dimension the persona asked
        # for before its preference was ever read.
        self.dims = [c for c in cols if c["semantic_role"] in ("categorical_dim", "state_flag")
                     and not c.get("high_cardinality")]
        self.big_dims = [c for c in cols if c["semantic_role"] == "categorical_dim"
                         and c.get("high_cardinality")][:2]
        self.labels = [c for c in cols if c["semantic_role"] == "entity_label"]
        fact_pk = set(self.fact.get("grain") or [])
        keys = [c for c in cols if c["semantic_role"] == "entity_key"
                and c["name"] not in fact_pk]
        rows_n = self.fact["row_count"] or 1
        grouping = [c for c in keys if (c.get("distinct") or rows_n) < rows_n * 0.5]
        # Cardinality alone picks the wrong entity often enough to be worth a knob:
        # a sales persona concentrates on customers, not postal codes. Mirrors
        # prefer_slices -- named first, everything else by cardinality after.
        wanted = {name: i for i, name in enumerate(plan.get("prefer_entities", []))}
        self.entities = sorted(grouping or keys,
                               key=lambda c: (wanted.get(c["name"], len(wanted)),
                                              c.get("distinct") or 0)) or keys
        # Join graph, verified edges only. A slice is worth offering only if the
        # data confirmed the relationship.
        self.tables = {t["name"]: t for t in card["tables"]}
        self.edges = [e for e in card.get("joins", []) if e.get("verified")]
        self.paths = self._reachable()
        for table_name, path in self.paths.items():
            table = self.tables[table_name]
            for column in table["columns"]:
                if (column["semantic_role"] in ("categorical_dim", "state_flag")
                        and not column.get("high_cardinality")):
                    self.dims.append({**column, "_table": table_name, "_path": path})
                if (column["semantic_role"] == "entity_label"
                        and table["role"] in ("dimension_like", "lookup")):
                    self.labels.append({**column, "_table": table_name, "_path": path})
        for column in cols:
            column.setdefault("_table", self.fact["name"])
            column.setdefault("_path", [])
        # Persona-declared composite measures the card cannot infer (net of a
        # discount, for instance). Fact-table columns only.
        for offset, derived in enumerate(plan.get("derived_measures", [])):
            path = []
            for required in derived.get("requires_tables", []):
                path += self.paths.get(required, [])
            entry = {
                "name": derived["name"], "_expr": derived["expr"],
                "_kind": derived.get("kind"),
                "_filter": derived.get("filter"), "_kpi_only": derived.get("kpi_only", False),
                "additivity": derived.get("additivity", "additive"),
                "semantic_role": "derived", "_table": self.fact["name"], "_path": path}
            # A derived rate is still a rate: it must be averaged, never totalled.
            # Routing it into self.measures would emit SUM(a/b) -- the exact
            # aggregation the additivity matrix forbids.
            if entry["additivity"] == "non_additive":
                self.rates.insert(0, entry)
                del self.rates[2:]
            else:
                self.measures.insert(offset, entry)
        shown = plan.get("show_measures")
        if shown:
            self.measures = [c for c in self.measures if c["name"] in shown] or self.measures
        preferred = {name: i for i, name in enumerate(plan.get("prefer_slices", []))}

        def dim_rank(column):
            if column["name"] in preferred:
                return (-1, preferred[column["name"]], 0, 0)
            owner = self.tables.get(column.get("_table", self.fact["name"]), {})
            on_dimension = owner.get("role") in ("dimension_like", "lookup")
            distinct = column.get("distinct") or 0
            useful_width = 0 if 3 <= distinct <= 60 else 1
            return (0 if on_dimension else 1, useful_width,
                    0 if column["semantic_role"] == "categorical_dim" else 1, distinct)

        prefer_time = plan.get("prefer_time", [])
        if prefer_time:
            candidates = [c for c in cols if c["semantic_role"] == "temporal_event"]
            for table_name, path in self.paths.items():
                for column in self.tables[table_name]["columns"]:
                    if column["semantic_role"] == "temporal_event":
                        candidates.append({**column, "_table": table_name, "_path": path})
            candidates.sort(key=lambda c: prefer_time.index(c["name"])
                            if c["name"] in prefer_time else 99)
            if candidates and candidates[0]["name"] in prefer_time:
                self.time = candidates[0]

        self.dims.sort(key=dim_rank)
        del self.dims[6:]  # now that preference has been applied, keep the top six
        self.labels.sort(key=lambda c: preferred.get(c["name"], 99))
        role_rank = {"derived": 0, "monetary_amount": 1, "quantity": 2}
        self.measures.sort(key=lambda c: role_rank.get(c["semantic_role"], 3))
        self.filters = self.dims[:2]
        self.grain = {"day": "DD", "week": "IW", "month": "MM",
                      "hour": "HH", "native": "MM"}.get(plan.get("grain", "month"), "MM")
        self.grain_label = plan.get("grain", "month")

    def _reachable(self, max_hops: int = 4) -> dict:
        """Shortest verified join path from the fact table to each other table."""
        found, frontier = {}, [(self.fact["name"], [])]
        while frontier:
            table_name, path = frontier.pop(0)
            if len(path) >= max_hops:
                continue
            for edge in self.edges:
                for src, dst, on_from, on_to in (
                        (edge["from_table"], edge["to_table"], edge["from_column"], edge["to_column"]),
                        (edge["to_table"], edge["from_table"], edge["to_column"], edge["from_column"])):
                    if src != table_name or dst == self.fact["name"] or dst in found:
                        continue
                    step = path + [{"table": dst, "left": src, "left_col": on_from,
                                    "right_col": on_to,
                                    "composite_on": edge.get("composite_on")}]
                    found[dst] = step
                    frontier.append((dst, step))
        return found

    def expr(self, column: dict) -> str:
        if column.get("_expr"):
            if column.get("_filter"):
                return f'CASE WHEN {column["_filter"]} THEN {column["_expr"]} ELSE 0 END'
            return column["_expr"]
        table = column.get("_table", self.fact["name"])
        cast = column.get("cast_expression")
        if cast:
            return cast.replace(q(column["name"]), f'{q(table)}.{q(column["name"])}')
        return f'{q(table)}.{q(column["name"])}'

    def window(self) -> str:
        """Restrict a breakdown to the same trailing window the KPIs report."""
        if not self.time:
            return ""
        return (f"  AND {self.expr(self.time)} > ADD_MONTHS("
                f"(SELECT MAX({self.expr(self.time)}) FROM {self.from_clause([self.time])}),"
                f" -12)\n")

    def series_measures(self) -> list:
        return [c for c in self.measures if not c.get("_kpi_only")]

    def source(self) -> str:
        return f'{q(self.schema)}.{q(self.fact["name"])}'

    def from_clause(self, columns: list[dict]) -> str:
        """FROM the fact, joining only the tables the chosen columns actually need."""
        needed, seen = [], set()
        for column in columns:
            for step in column.get("_path", []):
                key = step["table"]
                if key not in seen:
                    seen.add(key)
                    needed.append(step)
        text = f'{self.source()} {q(self.fact["name"])}'
        for step in needed:
            pairs = step.get("composite_on") or [{"from": step["left_col"],
                                                  "to": step["right_col"]}]
            on = " AND ".join(f'{q(step["table"])}.{q(pair["to"])}'
                              f' = {q(step["left"])}.{q(pair["from"])}' for pair in pairs)
            text += f'\n  JOIN {q(self.schema)}.{q(step["table"])} {q(step["table"])} ON {on}'
        return text

    def where(self, require_time: bool = False) -> str:
        clauses = []
        if require_time and self.time:
            clauses.append(f"  AND {self.expr(self.time)} IS NOT NULL")
        for index, column in enumerate(self.filters):
            token = f"f{index}"
            clauses.append(f"  AND ({{{token}!s}} IS NULL OR "
                           f"{self.expr(column)} = {{{token}!s}})")
        return "\n".join(clauses)

    def params(self) -> list[str]:
        return [f"f{i}" for i in range(len(self.filters))]


# ------------------------------------------------------------------ SQL
def sql_kpi(m: Model) -> str:
    parts = [f"-- Headline totals for the current window and the one before it.",
             "-- Aggregations follow each measure's additivity class; nothing here sums a rate."]
    if m.time:
        parts.append(f"WITH bounds AS (SELECT MAX({m.expr(m.time)}) AS D "
                     f"FROM {m.from_clause([m.time])})")
        sel = ["  (SELECT CAST(D AS VARCHAR(32)) FROM bounds) AS ASOF"]
        for i, meas in enumerate(m.measures):
            e = m.expr(meas)
            sel.append(f"  SUM(CASE WHEN {m.expr(m.time)} > ADD_MONTHS(b.D, -12) "
                       f"THEN {e} ELSE 0 END) AS M{i}_CUR")
            sel.append(f"  SUM(CASE WHEN {m.expr(m.time)} <= ADD_MONTHS(b.D, -12) "
                       f"AND {m.expr(m.time)} > ADD_MONTHS(b.D, -24) THEN {e} ELSE 0 END) AS M{i}_PRI")
        for i, rate in enumerate(m.rates):
            sel.append(f"  AVG(CASE WHEN {m.expr(m.time)} > ADD_MONTHS(b.D, -12) "
                       f"THEN {m.expr(rate)} END) AS R{i}_CUR")
            sel.append(f"  AVG(CASE WHEN {m.expr(m.time)} <= ADD_MONTHS(b.D, -12) "
                       f"AND {m.expr(m.time)} > ADD_MONTHS(b.D, -24) THEN {m.expr(rate)} END) AS R{i}_PRI")
        sel.append("  COUNT(*) AS ROWS_N")
        if m.entities:
            sel.append(f"  COUNT(DISTINCT {m.expr(m.entities[0])}) AS ENTITIES_N")
        parts.append("SELECT\n" + ",\n".join(sel))
        parts.append(f"FROM {m.from_clause(m.measures + m.rates + m.filters)} "
                     f"CROSS JOIN bounds b")
        parts.append("WHERE 1=1\n" + m.where())
    else:
        sel = ["  CAST(NULL AS VARCHAR(32)) AS ASOF"]
        for i, meas in enumerate(m.measures):
            sel.append(f"  SUM({m.expr(meas)}) AS M{i}_CUR")
            sel.append(f"  CAST(NULL AS DOUBLE) AS M{i}_PRI")
        for i, rate in enumerate(m.rates):
            sel.append(f"  AVG({m.expr(rate)}) AS R{i}_CUR")
            sel.append(f"  CAST(NULL AS DOUBLE) AS R{i}_PRI")
        sel.append("  COUNT(*) AS ROWS_N")
        if m.entities:
            sel.append(f"  COUNT(DISTINCT {m.expr(m.entities[0])}) AS ENTITIES_N")
        parts.append("SELECT\n" + ",\n".join(sel))
        parts.append(f"FROM {m.from_clause(m.measures + m.rates + m.filters)}")
        parts.append("WHERE 1=1\n" + m.where())
    return "\n".join(parts) + "\n"


def sql_trend(m: Model) -> str:
    if not m.time:
        return ""
    sel = [f"  TO_CHAR(TRUNC({m.expr(m.time)}, '{m.grain}'), 'YYYY-MM-DD') AS PERIOD"]
    for i, meas in enumerate(m.series_measures()):
        sel.append(f"  SUM({m.expr(meas)}) AS M{i}")
    return ("-- Measures over the persona's cadence. One row per period.\n"
            "SELECT\n" + ",\n".join(sel) +
            f"\nFROM {m.from_clause(m.series_measures() + m.filters + [m.time])}"
            f"\nWHERE TRUNC({m.expr(m.time)}, '{m.grain}') < "
            f"(SELECT TRUNC(MAX({m.expr(m.time)}), '{m.grain}') "
            f"FROM {m.from_clause([m.time])})\n"
            + m.where(require_time=True) +
            f"\nGROUP BY TRUNC({m.expr(m.time)}, '{m.grain}')\nORDER BY 1\n")


def sql_composition(m: Model) -> str:
    if not m.dims or not m.measures:
        return ""
    dim = m.dims[0]
    sel = [f"  {m.expr(dim)} AS SLICE"]
    for i, meas in enumerate(m.series_measures()):
        sel.append(f"  SUM({m.expr(meas)}) AS M{i}")
    sel.append("  COUNT(*) AS ROWS_N")
    return (f"-- {m.measures[0]['name']} composed by {dim['name']}.\n"
            "SELECT\n" + ",\n".join(sel) +
            f"\nFROM {m.from_clause(m.series_measures() + m.filters + [dim])}"
            f"\nWHERE {m.expr(dim)} IS NOT NULL\n" + m.window() + m.where() +
            f"\nGROUP BY {m.expr(dim)}\nORDER BY 2 DESC\n")


def sql_worklist(m: Model) -> str:
    if not m.measures:
        return ""
    key = (m.labels or m.entities or m.dims or m.measures)[0]
    sel = [f"  {m.expr(key)} AS ENTITY"]
    for i, meas in enumerate(m.series_measures()):
        sel.append(f"  SUM({m.expr(meas)}) AS M{i}")
    sel.append("  COUNT(*) AS ROWS_N")
    return (f"-- Ranked {key['name']}. Bounded: a worklist is never unbounded.\n"
            "SELECT\n" + ",\n".join(sel) +
            f"\nFROM {m.from_clause(m.series_measures() + m.filters + [key])}"
            f"\nWHERE {m.expr(key)} IS NOT NULL\n" + m.window() + m.where() +
            f"\nGROUP BY {m.expr(key)}\nORDER BY 2 DESC\nLIMIT 25\n")


def sql_crosstab(m: Model) -> str:
    """Two dimensions at once: where a total actually comes from."""
    if len(m.dims) < 2 or not m.series_measures():
        return ""
    a, b = m.dims[0], m.dims[1]
    sel = [f"  {m.expr(a)} AS DIM_A", f"  {m.expr(b)} AS DIM_B"]
    for i, meas in enumerate(m.series_measures()):
        sel.append(f"  SUM({m.expr(meas)}) AS M{i}")
    return (f"-- {m.series_measures()[0]['name']} across {a['name']} x {b['name']}.\n"
            "SELECT\n" + ",\n".join(sel) +
            f"\nFROM {m.from_clause(m.series_measures() + m.filters + [a, b])}"
            f"\nWHERE {m.expr(a)} IS NOT NULL AND {m.expr(b)} IS NOT NULL\n"
            + m.window() + m.where() +
            f"\nGROUP BY {m.expr(a)}, {m.expr(b)}\nORDER BY 3 DESC\n")


def sql_concentration(m: Model) -> str:
    """How much of the total sits with the few largest entities."""
    key = (m.labels or m.entities or m.dims)
    if not key or not m.series_measures():
        return ""
    key = key[0]
    measure = m.series_measures()[0]
    return (f"-- Share of {measure['name']} held by each entity, largest first.\n"
            f"SELECT\n  {m.expr(key)} AS ENTITY,\n  SUM({m.expr(measure)}) AS SHARE_VALUE\n"
            f"FROM {m.from_clause([measure] + m.filters + [key])}"
            f"\nWHERE {m.expr(key)} IS NOT NULL\n" + m.window() + m.where() +
            f"\nGROUP BY {m.expr(key)}\nORDER BY 2 DESC\n")


def sql_filters(m: Model) -> str:
    if not m.filters:
        return "SELECT 'NONE' AS DIM, 'NONE' AS DIM_VALUE FROM DUAL WHERE 1=0\n"
    unions = [f"SELECT '{c['name']}' AS DIM, CAST({m.expr(c)} AS VARCHAR(200)) AS DIM_VALUE\n"
              f"FROM {m.from_clause([c])} WHERE {m.expr(c)} IS NOT NULL\n"
              f"GROUP BY {m.expr(c)}" for c in m.filters]
    return ("-- Filter values read live, never hardcoded.\n"
            + "\nUNION ALL\n".join(unions) + "\nORDER BY 1, 2\n")




# --------------------------------------------------------------- llm_sql.py
LLM_MODULE = '''"""Optional LLM-backed text-to-SQL, with a guard the model cannot talk past.

The key is resolved in two steps, environment first:

  1. ANTHROPIC_API_KEY in the process environment.
  2. ~/.exasol-starter-kit/credentials/anthropic_api_key, owner-readable only.

Step 2 exists because dash-server is started by a launchd boot entry with no
EnvironmentVariables, so a key exported in a login shell never reaches it. The
kit already resolves the database password this way. The key is never written
to a file by this module, never logged, and never placed in app source.

With no key present the caller falls back to the deterministic template engine
and the panel still works, so this module is optional by construction.

Every candidate query, whoever wrote it, must survive the guard below before it
reaches the database. The database user is read-only, which is the floor; the
guard is what stops expensive, sprawling or off-schema queries above it.
"""
import os, re, stat

MODEL = "claude-opus-5"
KEY_FILE = os.path.join(os.path.expanduser("~"), ".exasol-starter-kit",
                        "credentials", "anthropic_api_key")
FORBIDDEN = re.compile(
    r"\\b(insert|update|delete|merge|drop|create|alter|truncate|grant|revoke|"
    r"commit|rollback|call|execute|import|export)\\b", re.I)

SYSTEM = (
    "You translate a business question into exactly one Exasol SELECT statement. "
    "Return SQL only: no prose, no code fence, no semicolon, no comments."
)


def read_key():
    """(key, error). Never returns the key in the error, and never logs it."""
    key = os.environ.get("ANTHROPIC_API_KEY")
    if key:
        return key.strip(), None
    try:
        info = os.stat(KEY_FILE)
    except OSError:
        return None, None                      # not configured; not an error
    # POSIX mode bits only. Python on Windows reports st_mode as 0o666 for every
    # regular file, so this guard would reject every key there and chmod could
    # not clear it -- Windows protects files with ACLs, not mode bits.
    if os.name != "nt" and info.st_mode & (stat.S_IRGRP | stat.S_IROTH):
        return None, (f"{KEY_FILE} is readable by other users. "
                      f"Run: chmod 600 {KEY_FILE}")
    try:
        with open(KEY_FILE, encoding="utf-8") as handle:
            key = handle.read().strip()
    except OSError as exc:
        return None, f"could not read the key file: {exc.strerror}"
    return (key, None) if key else (None, "the key file is empty")


def guard(sql, schema, table, row_cap=500):
    """Return (safe_sql, error). Rejects anything that is not one bounded SELECT."""
    if not sql or not sql.strip():
        return None, "empty query"
    text = sql.strip().rstrip(";").strip()
    if ";" in text:
        return None, "rejected: more than one statement"
    if "--" in text or "/*" in text:
        return None, "rejected: comments can hide a second statement"
    if not re.match(r"^(select|with)\\b", text, re.I):
        return None, "rejected: only SELECT is allowed"
    if FORBIDDEN.search(text):
        return None, "rejected: contains a data- or schema-modifying keyword"
    # FROM is also a separator inside EXTRACT/SUBSTRING/TRIM/OVERLAY/POSITION,
    # where it introduces no table at all: EXTRACT(YEAR FROM "T"."C") would
    # otherwise be read as schema "T", table "C" and refused. Blank the keyword
    # out for the scan only; `text` is untouched and is what actually runs.
    scan = re.sub(r'\\b(?:EXTRACT|SUBSTRING|TRIM|OVERLAY|POSITION)\\s*\\((?:[^()]|\\([^()]*\\))*\\)',
                  lambda m: re.sub(r'\\bFROM\\b', '____', m.group(0), flags=re.I),
                  text, flags=re.I | re.S)
    sources = re.findall(r'\\b(?:from|join)\\s+"([A-Za-z0-9_ ]+)"\\s*\\.\\s*"([A-Za-z0-9_ ]+)"',
                         scan, re.I)
    unqualified = re.findall(r'\\b(?:from|join)\\s+(?!")([A-Za-z0-9_]+)', scan, re.I)
    if not sources:
        return None, "rejected: every table must be written as \\"SCHEMA\\".\\"TABLE\\""
    for ref_schema, _ in sources:
        if ref_schema.upper() != schema.upper():
            return None, f"rejected: reads from schema {ref_schema}, outside this dashboard"
    if unqualified:
        return None, (f"rejected: unqualified table reference "
                      f"{unqualified[0]}; every table must name its schema")
    return f"SELECT * FROM ({text}) LIMIT {row_cap}", None


def propose_sql(question, schema, table, columns, joins=None, as_of=None):
    """Ask the model for SQL. Returns (sql, error). Not configured -> (None, None)."""
    key, key_error = read_key()
    if key_error:
        return None, key_error
    if not key:
        return None, None
    try:
        import anthropic
    except ImportError:
        return None, ("the anthropic package is not installed in this app "
                      "environment; add it to requirements.txt and redeploy")

    catalog = "\\n".join(
        f'  "{c.get("table", table)}"."{c["name"]}"  {c["role"]}  {c["additivity"]}'
        for c in columns)
    join_text = ("\\nAvailable joins from the fact table (use only these):\\n  "
                 + "\\n  ".join(joins)) if joins else ""
    prompt = (
        f'Fact table: "{schema}"."{table}" (alias it as "{table}").\\n'
        f"Columns (table.name, semantic role, additivity):\\n"
        f"{catalog}{join_text}\\n\\nRules:\\n"
        f'- Schema-qualify every table as "{schema}"."TABLE", aliased to its own name.\\n'
        f"- Quote every identifier with double quotes.\\n"
        f"- NEVER SUM a column whose additivity is non_additive, attribute or "
        f"count_only. Average those instead.\\n"
        f"- A semi_additive column may not be summed across time.\\n"
        + (f"- This dataset ends on {as_of}. It is historical: CURRENT_DATE and "
           f"NOW() fall outside it and match no rows. Resolve 'last year', "
           f"'recent' and similar against {as_of}, with literal dates or "
           f"ADD_MONTHS(DATE '{as_of}', -n).\\n" if as_of else "")
        + f"\\nQuestion: {question}")

    client = anthropic.Anthropic(api_key=key)
    try:
        response = client.beta.messages.create(
            model=MODEL,
            max_tokens=16000,
            system=SYSTEM,
            betas=["server-side-fallback-2026-07-01"],
            fallbacks="default",
            messages=[{"role": "user", "content": prompt}],
        )
    except anthropic.RateLimitError:
        return None, "the model is rate limited; try again shortly"
    except anthropic.AuthenticationError:
        return None, "the API key was rejected"
    except anthropic.APIStatusError as exc:
        return None, f"model call failed (HTTP {exc.status_code})"
    except anthropic.APIConnectionError:
        return None, "could not reach the API; check the network"

    # A refusal returns HTTP 200 with no usable content: check before reading it.
    if getattr(response, "stop_reason", None) == "refusal":
        return None, "the model declined to answer this question"
    text = "".join(block.text for block in response.content
                   if getattr(block, "type", None) == "text")
    text = re.sub(r"^```(?:sql)?|```$", "", text.strip(), flags=re.M).strip()
    return (text, None) if text else (None, "the model returned no SQL")
'''


APP_TEMPLATE = Template("""\"\"\"$title

Generated by persona-metrics stage 6 from the schema card for $schema.
Persona: $persona  ·  Grain: $grain_label
\"\"\"
import importlib.util, io, csv, re
from pathlib import Path

import plotly.graph_objects as go
from dash import Dash, Input, Output, State, dcc, html, no_update

_SPEC = importlib.util.spec_from_file_location(
    "gen_exasol_helper", Path(__file__).with_name("dash_server_exasol.py"))
_MOD = importlib.util.module_from_spec(_SPEC); _SPEC.loader.exec_module(_MOD)
load_row, load_rows = _MOD.load_row, _MOD.load_rows
has_error, render_error_panel = _MOD.has_error, _MOD.render_error_panel
_LSPEC = importlib.util.spec_from_file_location(
    "gen_llm_sql", Path(__file__).with_name("llm_sql.py"))
_LLM = importlib.util.module_from_spec(_LSPEC); _LSPEC.loader.exec_module(_LLM)

# --- palette: validated categorical hues on a warm neutral ground -------------
INK, INK_2, MUTED = "#12100e", "#4b463f", "#8a837a"
SURFACE, PAGE = "#ffffff", "#f4f1ec"
LINE, HAIR = "#e6e1d8", "#f0ece4"
SERIES = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100", "#7c5cd6"]
TINT = ["#eaf2fd", "#fdefe9", "#e8f7f1", "#fdf5e3", "#f1ecfd"]
GOOD, GOOD_BG = "#0a7d4b", "#e6f5ee"
BAD, BAD_BG = "#c0392b", "#fdecea"
WARN, WARN_BG = "#a8730a", "#fdf3e0"
FONT = '"Plus Jakarta Sans", system-ui, -apple-system, "Segoe UI", sans-serif'
JAKARTA = ("https://fonts.googleapis.com/css2?"
           "family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap")

CARD = {"backgroundColor": SURFACE, "border": "1px solid " + LINE,
        "borderRadius": "14px", "padding": "1.15rem 1.3rem",
        "boxShadow": "0 1px 2px rgba(18,16,14,0.04), 0 8px 24px -16px rgba(18,16,14,0.18)"}
H2 = {"fontSize": "13px", "margin": 0, "color": INK, "fontWeight": 700,
      "letterSpacing": "0.02em", "textTransform": "uppercase"}

MEASURES = $measures
SERIES_MEASURES = $series_measures
KINDS = $kinds
RATES = $rates
MEASURE_SQL = $measure_sql
DIMS = $dims
CHAT_DIMS = $chat_dims
VALUES = $values
FILTERS = $filters
SCHEMA, TABLE = $schema_r, $table_r
HAS_TIME = $has_time
TIME_EXPR = $time_expr
MODEL_READY = bool(_LLM.read_key()[0])
SETUP_KEY_CMD = $setup_key_cmd
ENTITY_NOUN = $entity_noun
REFUSED = $refused
NEGATIVES = $negatives
CUT_METRICS = $cut
CATALOG = $catalog
JOINS = $joins
_SNAPSHOT = {}


def pretty(name):
    \"\"\"Turn a warehouse column name into something a board will read.\"\"\"
    text = re.sub(r"^[A-Z]{1,3}_", "", str(name))
    text = re.sub(r"[_\\-]+", " ", text).strip()
    if text.isupper() or text.islower():
        text = text.title()
    return text


def _f(v, d=0.0):
    try: return float(v)
    except (TypeError, ValueError): return d


def fmt(value, measure=None):
    a = _f(value)
    money = KINDS.get(measure) == "money"
    prefix = "$$" if money else ""
    for cut, suffix in ((1e9, "B"), (1e6, "M"), (1e3, "K")):
        if abs(a) >= cut:
            return f"{prefix}{a/cut:,.1f}{suffix}"
    if money:
        return f"{prefix}{a:,.0f}"
    return f"{a:,.0f}" if a == int(a) else f"{a:,.2f}"


def fmt_rate(value, measure=None):
    \"\"\"Not every non-additive measure is a fraction.

    Formatting all of them as percentages turns an average of 4.24 days into
    "424.4%". Trust the measure's declared kind instead.
    \"\"\"
    kind = KINDS.get(measure)
    if kind == "percent":
        return f"{_f(value):.1%}"
    if kind == "money":
        return fmt(value, measure)
    return f"{_f(value):,.2f}"


def delta_pill(pct, good_when_up=True):
    up = pct >= 0
    good = up if good_when_up else not up
    colour, background = (GOOD, GOOD_BG) if good else (BAD, BAD_BG)
    return html.Span(f"{'▲' if up else '▼'} {abs(pct):,.1f}%",
                     style={"color": colour, "backgroundColor": background,
                            "fontSize": "11.5px", "fontWeight": 700,
                            "padding": "0.15rem 0.45rem", "borderRadius": "999px"})


def window_label(asof):
    \"\"\"The period a KPI covers, on the KPI itself.

    Every headline is a trailing twelve months, but only the page header said so.
    A reader who totals the same measure elsewhere gets a different number and no
    way to see why.
    \"\"\"
    if not HAS_TIME or not asof:
        return None
    text = str(asof)[:10]
    try:
        year, month, day = (int(part) for part in text.split("-"))
    except ValueError:
        return f"12 months to {text}"
    names = ("Jan", "Feb", "Mar", "Apr", "May", "Jun",
             "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")
    start_month, start_year = month + 1, year - 1
    if start_month > 12:
        start_month, start_year = start_month - 12, start_year + 1
    return f"{names[start_month - 1]} {start_year} - {names[month - 1]} {year}"


def kpi_tile(label, value, index, delta=None, note=None, good_when_up=True, period=None):
    accent = SERIES[index % len(SERIES)]
    body = [html.Div(label, style={"fontSize": "11.5px", "color": MUTED, "fontWeight": 600,
                                   "letterSpacing": "0.04em", "textTransform": "uppercase"})]
    if period:
        body.append(html.Div(period, style={"fontSize": "10.5px", "color": MUTED,
                                            "marginTop": "0.15rem", "opacity": 0.85}))
    body.append(html.Div(value, style={"fontSize": "31px", "fontWeight": 700, "color": INK,
                                       "margin": "0.35rem 0 0.45rem",
                                       "letterSpacing": "-0.02em"}))
    row = []
    if delta is not None:
        row.append(delta_pill(delta, good_when_up))
    if note:
        row.append(html.Span(note, style={"fontSize": "11.5px", "color": MUTED,
                                          "marginLeft": "0.4rem" if row else 0}))
    if row:
        body.append(html.Div(row, style={"display": "flex", "alignItems": "center",
                                         "flexWrap": "wrap", "gap": "0.15rem"}))
    return html.Div([html.Div(style={"height": "3px", "backgroundColor": accent,
                                     "borderRadius": "999px", "width": "34px",
                                     "marginBottom": "0.75rem"}),
                     html.Div(body)], style=CARD)


def insight_card(finding):
    \"\"\"One finding, with its meaning and its action attached rather than split
    across parallel columns that never balance.\"\"\"
    tone = finding["tone"]
    colour, background, icon = {"good": (GOOD, GOOD_BG, "▲"), "bad": (BAD, BAD_BG, "▼"),
                                "warn": (WARN, WARN_BG, "!"),
                                "info": (INK_2, PAGE, "•")}[tone]
    body = [html.Div(finding["text"], style={"fontSize": "13.5px", "color": INK,
                                             "fontWeight": 600, "lineHeight": 1.45})]
    if finding.get("meaning"):
        body.append(html.Div(finding["meaning"],
                             style={"fontSize": "12.5px", "color": INK_2,
                                    "lineHeight": 1.5, "marginTop": "0.25rem"}))
    if finding.get("action"):
        body.append(html.Div([
            html.Span("ACTION", style={"fontSize": "9.5px", "fontWeight": 800,
                                       "letterSpacing": "0.08em", "color": colour,
                                       "marginRight": "0.45rem"}),
            html.Span(finding["action"], style={"fontSize": "12.5px", "color": INK_2}),
        ], style={"marginTop": "0.45rem", "paddingTop": "0.45rem",
                  "borderTop": "1px dashed " + LINE}))
    return html.Div([
        html.Div(icon, style={"color": colour, "backgroundColor": background,
                              "fontWeight": 800, "fontSize": "12px", "minWidth": "24px",
                              "height": "24px", "borderRadius": "8px", "display": "flex",
                              "alignItems": "center", "justifyContent": "center",
                              "flexShrink": 0}),
        html.Div(body),
    ], style={"display": "flex", "gap": "0.7rem", "alignItems": "flex-start",
              "padding": "0.75rem 0.85rem", "borderRadius": "12px",
              "backgroundColor": SURFACE, "border": "1px solid " + LINE,
              "borderLeft": f"3px solid {colour}"})


def base_fig(title, y_title, height=340):
    figure = go.Figure()
    figure.update_layout(
        title={"text": title, "font": {"size": 13, "color": INK, "family": FONT},
               "x": 0, "xanchor": "left", "y": 0.95},
        height=height, margin={"l": 62, "r": 26, "t": 52, "b": 46},
        paper_bgcolor=SURFACE, plot_bgcolor=SURFACE,
        font={"family": FONT, "color": INK_2, "size": 12}, hovermode="x unified",
        legend={"orientation": "h", "yanchor": "bottom", "y": 1.0, "x": 0,
                "font": {"size": 11}, "bgcolor": "rgba(0,0,0,0)"},
        xaxis={"showgrid": False, "linecolor": LINE, "tickfont": {"color": MUTED},
               "ticks": "outside", "tickcolor": LINE},
        yaxis={"title": {"text": y_title, "font": {"size": 11, "color": MUTED}},
               "gridcolor": HAIR, "zerolinecolor": LINE, "linecolor": "rgba(0,0,0,0)",
               "tickfont": {"color": MUTED}})
    return figure


def empty_fig(message, height=340):
    figure = base_fig("", "", height)
    figure.add_annotation(text=message, showarrow=False, font={"color": MUTED, "size": 12})
    figure.update_xaxes(visible=False); figure.update_yaxes(visible=False)
    return figure


MONEY_WORDS = ("sales", "revenue", "price", "cost", "amount", "profit", "spend",
               "discount given", "margin", "value", "total")


def looks_monetary(column_name):
    text = str(column_name).lower().replace("_", " ")
    return any(word in text for word in MONEY_WORDS)


def cell_value(column_name, value):
    \"\"\"Render a raw query value for display: money as a currency amount.\"\"\"
    if isinstance(value, bool) or value is None:
        return "" if value is None else str(value)
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if looks_monetary(column_name):
        for cut, suffix in ((1e9, "B"), (1e6, "M"), (1e3, "K")):
            if abs(number) >= cut:
                return f"$${number/cut:,.1f}{suffix}"
        return f"$${number:,.2f}"
    if abs(number) >= 1000:
        return f"{number:,.0f}"
    return f"{number:,.4f}".rstrip("0").rstrip(".")


def data_table(rows, columns, aligns=None, raw=False):
    aligns = aligns or (["left"] + ["right"] * (len(columns) - 1))
    head = html.Thead(html.Tr([
        html.Th(c, style={"textAlign": a, "padding": "0.55rem 0.75rem",
                          "borderBottom": "2px solid " + LINE, "fontSize": "11px",
                          "color": MUTED, "fontWeight": 700, "letterSpacing": "0.04em",
                          "textTransform": "uppercase", "whiteSpace": "nowrap"})
        for c, a in zip(columns, aligns)]))
    body = html.Tbody([
        html.Tr([html.Td(str(r.get(c, "")) if raw else cell_value(c, r.get(c)), style={
            "textAlign": a, "padding": "0.5rem 0.75rem", "borderBottom": "1px solid " + HAIR,
            "fontSize": "13px", "color": INK, "fontVariantNumeric": "tabular-nums",
            "whiteSpace": "nowrap"}) for c, a in zip(columns, aligns)],
            style={"backgroundColor": SURFACE if i % 2 == 0 else PAGE})
        for i, r in enumerate(rows)])
    return html.Div(html.Table([head, body], style={"borderCollapse": "collapse",
                    "width": "100%", "minWidth": "560px"}), style={"overflowX": "auto"})


def _legacy_build_insights(kpi, trend, comp):
    \"\"\"Rule-based. Every figure comes from a query; none of this is generated prose.\"\"\"
    out = []
    for i, name in enumerate(MEASURES):
        cur, pri = _f(kpi.get(f"M{i}_CUR")), _f(kpi.get(f"M{i}_PRI"))
        if pri and abs(cur / pri - 1) * 100 >= 5:
            change = (cur / pri - 1) * 100
            out.append((f"{pretty(name)} is {abs(change):.0f}% "
                        f"{'below' if change < 0 else 'above'} the prior period "
                        f"({fmt(cur, name)} vs {fmt(pri, name)}).",
                        "bad" if change < 0 else "good"))
    if comp and len(comp) > 1:
        total = sum(_f(r.get("M0")) for r in comp) or 1
        top = comp[0]
        share = _f(top.get("M0")) / total * 100
        label = pretty(top.get("SLICE"))
        headline = SERIES_MEASURES[0] if SERIES_MEASURES else (MEASURES[0] if MEASURES else "")
        if share >= 40:
            out.append((f"{label} alone carries {share:.0f}% of {pretty(headline)} "
                        f"({fmt(top.get('M0'), headline)}) — concentration worth watching.",
                        "warn"))
        else:
            out.append((f"Largest group {label} holds {share:.0f}% of "
                        f"{pretty(headline)}; spread across {len(comp)} groups.", "info"))
    if trend and len(trend) >= 4:
        values = [_f(r.get("M0")) for r in trend]
        recent, earlier = values[-3:], values[-6:-3] or values[:3]
        if sum(earlier):
            drift = (sum(recent) / len(recent)) / (sum(earlier) / len(earlier)) - 1
            if abs(drift) >= 0.08:
                out.append((f"The last three periods average {abs(drift)*100:.0f}% "
                            f"{'below' if drift < 0 else 'above'} the three before them — "
                            f"a sustained move, not a single-period wobble.",
                            "bad" if drift < 0 else "good"))
    for cut in CUT_METRICS:
        out.append((f"{cut} was derived and then dropped: flat across every dimension "
                    f"tested, so charting it would imply a signal that is not there.", "info"))
    if not out:
        out.append(("Nothing outside normal range this period.", "info"))
    return out


def make_figures(kpi, trend, comp, conc, headline_measure):
    \"\"\"Build every chart once, so the page and the shared file cannot drift.\"\"\"
    cfig = base_fig(f"{pretty(SERIES_MEASURES[0]) if SERIES_MEASURES else 'Rows'} by "
                    f"{pretty(DIMS[0]) if DIMS else 'group'}",
                    pretty(SERIES_MEASURES[0]) if SERIES_MEASURES else "rows", 330)
    if comp and not has_error(comp):
        top = list(reversed(comp[:12]))
        labels = [pretty(r.get("SLICE")) for r in top]
        values_ = [_f(r.get("M0")) for r in top]
        cfig.add_trace(go.Bar(y=labels, x=values_, orientation="h",
            marker={"color": SERIES[0], "line": {"color": SURFACE, "width": 2}},
            text=[fmt(v, SERIES_MEASURES[0] if SERIES_MEASURES else None)
                  for v in values_],
            textposition="outside", textfont={"color": INK_2, "size": 11},
            hovertemplate="%{y} — %{x:,.0f}<extra></extra>"))
        cfig.update_layout(bargap=0.32, showlegend=False, hovermode="closest",
                           margin={"l": 130, "r": 90, "t": 52, "b": 40})
        cfig.update_xaxes(showgrid=True, gridcolor=HAIR)
        cfig.update_yaxes(showgrid=False, title=None)

    if (HAS_TIME and trend and not has_error(trend) and len(SERIES_MEASURES) >= 2):
        numerator, denominator = SERIES_MEASURES[1], SERIES_MEASURES[0]
        rfig = base_fig(f"{pretty(numerator)} as % of {pretty(denominator)}",
                        "% of " + pretty(denominator), 300)
        periods = [r.get("PERIOD") for r in trend]
        ratio = [(_f(r.get("M1")) / _f(r.get("M0")) * 100) if _f(r.get("M0")) else None
                 for r in trend]
        clean_ratio = [v for v in ratio if v is not None]
        rfig.add_trace(go.Scatter(
            x=periods, y=ratio, mode="lines", name=pretty(numerator),
            line={"color": SERIES[2], "width": 2.4},
            hovertemplate="%{y:.2f}%<extra></extra>"))
        if clean_ratio:
            mean_ratio = sum(clean_ratio) / len(clean_ratio)
            rfig.add_hline(y=mean_ratio, line_dash="dot", line_color=MUTED,
                           annotation_text=f"mean {mean_ratio:.1f}%",
                           annotation_position="top left",
                           annotation_font={"size": 10, "color": MUTED})
            pad = max((max(clean_ratio) - min(clean_ratio)) * 0.25, 0.5)
            rfig.update_yaxes(range=[min(clean_ratio) - pad, max(clean_ratio) + pad])
        rfig.update_layout(showlegend=False)
    else:
        rfig = empty_fig("Needs two measures and a time column.", 300)

    kfig = base_fig(f"Concentration of {pretty(headline_measure)}",
                    "cumulative share of total", 300,)
    if conc and not has_error(conc):
        values = sorted((_f(r.get("SHARE_VALUE")) for r in conc), reverse=True)
        total = sum(values) or 1
        running, cumulative = 0.0, []
        for v in values:
            running += v
            cumulative.append(running / total * 100)
        cutoff = min(len(cumulative), 200)
        kfig.add_trace(go.Scatter(
            x=list(range(1, cutoff + 1)), y=cumulative[:cutoff], mode="lines",
            name="Cumulative share", line={"color": SERIES[4], "width": 2.4},
            fill="tozeroy", fillcolor="rgba(124,92,214,0.08)",
            hovertemplate="top %{x} hold %{y:.1f}%<extra></extra>"))
        if len(cumulative) >= 10:
            kfig.add_trace(go.Scatter(
                x=[10], y=[cumulative[9]], mode="markers+text", showlegend=False,
                marker={"color": SERIES[4], "size": 10,
                        "line": {"color": SURFACE, "width": 2}},
                text=[f"  top 10 = {cumulative[9]:.1f}%"], textposition="middle right",
                textfont={"color": INK_2, "size": 11}, hoverinfo="skip"))
        kfig.update_layout(showlegend=False, hovermode="closest",
                           margin={"l": 62, "r": 110, "t": 52, "b": 46})
        kfig.update_xaxes(title={"text": ENTITY_NOUN + ", largest first",
                                 "font": {"size": 11, "color": MUTED}})
    else:
        kfig = empty_fig("No entity breakdown available.", 300)

    tfig = None
    if HAS_TIME:
        tfig = base_fig(f"{pretty(SERIES_MEASURES[0]) if SERIES_MEASURES else 'Rows'}"
                        f" over time (complete periods only)",
                        pretty(SERIES_MEASURES[0]) if SERIES_MEASURES else "rows", 360)
        if trend and not has_error(trend):
            periods = [r.get("PERIOD") for r in trend]
            for i, name in enumerate(SERIES_MEASURES[:2]):
                series = [_f(r.get(f"M{i}")) for r in trend]
                tfig.add_trace(go.Scatter(
                    x=periods, y=series, mode="lines", name=pretty(name),
                    line={"color": SERIES[i], "width": 2.4, "shape": "spline",
                          "smoothing": 0.35},
                    fill="tozeroy" if i == 0 else None,
                    fillcolor="rgba(42,120,214,0.07)" if i == 0 else None,
                    hovertemplate=pretty(name) + " %{y:,.0f}<extra></extra>"))
            if periods:
                last = _f(trend[-1].get("M0"))
                tfig.add_trace(go.Scatter(
                    x=[periods[-1]], y=[last], mode="markers+text", showlegend=False,
                    marker={"color": SERIES[0], "size": 10,
                            "line": {"color": SURFACE, "width": 2}},
                    text=["  " + fmt(last, SERIES_MEASURES[0] if SERIES_MEASURES
                                     else None)],
                    textposition="middle right",
                    textfont={"color": INK_2, "size": 11}, hoverinfo="skip"))
                tfig.update_layout(margin={"l": 62, "r": 80, "t": 52, "b": 46})
    return {"composition": cfig, "rates": rfig, "concentration": kfig,
            "trend": tfig}


def share_html(title, caption, tiles, findings, figures, tables):
    \"\"\"A single self-contained HTML file: charts, insights and tables, no server.

    Plotly is inlined rather than pulled from a CDN, so the file still renders in
    five years, on a plane, from an email attachment.
    \"\"\"
    import plotly.io as pio

    tone_css = {"good": ("#0a7d4b", "#e6f5ee"), "bad": ("#c0392b", "#fdecea"),
                "warn": ("#a8730a", "#fdf3e0"), "info": ("#4b463f", "#f4f1ec")}

    def cards(group):
        if not group:
            return '<p class="muted">Nothing to report.</p>'
        out = []
        for finding in group:
            colour, background = tone_css.get(finding["tone"], tone_css["info"])
            meaning = (f'<div class="mean">{finding["meaning"]}</div>'
                       if finding.get("meaning") else "")
            action = (f'<div class="act"><span style="color:{colour}">ACTION</span> '
                      f'{finding["action"]}</div>' if finding.get("action") else "")
            out.append(f'<div class="ins" style="border-left:3px solid {colour}">'
                       f'<span class="dot" style="color:{colour};background:{background}">'
                       f'&#9632;</span><div><div class="head">{finding["text"]}</div>'
                       f'{meaning}{action}</div></div>')
        return "".join(out)

    tile_html = "".join(
        f'<div class="tile"><div class="accent" style="background:{SERIES[i % len(SERIES)]}">'
        f'</div><div class="lbl">{t["label"]}</div><div class="val">{t["value"]}</div>'
        f'<div class="delta" style="color:{"#0a7d4b" if t["good"] else "#c0392b"}">'
        f'{t["delta"]}</div><div class="note">{t["note"]}</div></div>'
        for i, t in enumerate(tiles))

    chart_html, first = [], True
    for key in ("trend", "composition", "rates", "concentration"):
        figure = figures.get(key)
        if figure is None:
            continue
        chart_html.append(
            '<div class="card">' +
            pio.to_html(figure, include_plotlyjs=(True if first else False),
                        full_html=False, config={"displayModeBar": False},
                        default_height="360px") + "</div>")
        first = False

    table_html = []
    for name, rows in tables:
        if not rows:
            continue
        columns = list(rows[0].keys())[:8]
        head = "".join(f"<th>{pretty(c)}</th>" for c in columns)
        body = "".join("<tr>" + "".join(f"<td>{cell_value(c, r.get(c))}</td>"
                                        for c in columns) + "</tr>" for r in rows)
        table_html.append(f'<div class="card"><h2>{name}</h2>'
                          f'<div class="scroll"><table><thead><tr>{head}</tr></thead>'
                          f"<tbody>{body}</tbody></table></div></div>")

    notes = "".join(f"<li>{n}</li>" for n in (NEGATIVES + REFUSED)) or \\
        "<li>Nothing was refused for this persona.</li>"

    return f\"\"\"<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title}</title>
<link href="{JAKARTA}" rel="stylesheet">
<style>
 :root {{ --ink:{INK}; --ink2:{INK_2}; --muted:{MUTED}; --line:{LINE}; --hair:{HAIR};
          --surface:{SURFACE}; --page:{PAGE}; }}
 * {{ box-sizing:border-box; }}
 body {{ font-family:{FONT}; background:var(--page); color:var(--ink);
         margin:0; padding:1.6rem; }}
 .wrap {{ max-width:1600px; margin:0 auto; }}
 header {{ padding:1.4rem 1.6rem; background:linear-gradient(180deg,var(--surface),var(--page));
           border:1px solid var(--line); border-radius:16px 16px 0 0; }}
 h1 {{ font-size:24px; margin:0; letter-spacing:-.02em; }}
 .cap {{ color:var(--muted); font-size:12.5px; margin-top:.3rem; }}
 .grid {{ display:grid; gap:.85rem; margin-top:.85rem; }}
 .kpis {{ grid-template-columns:repeat(auto-fit,minmax(196px,1fr)); }}
 .three {{ grid-template-columns:repeat(auto-fit,minmax(330px,1fr)); }}
 .two {{ grid-template-columns:repeat(auto-fit,minmax(330px,1fr)); }}
 .card,.tile {{ background:var(--surface); border:1px solid var(--line);
                border-radius:14px; padding:1.15rem 1.3rem;
                box-shadow:0 1px 2px rgba(18,16,14,.04),0 8px 24px -16px rgba(18,16,14,.18); }}
 .accent {{ height:3px; width:34px; border-radius:999px; margin-bottom:.75rem; }}
 .lbl {{ font-size:11.5px; color:var(--muted); font-weight:600; letter-spacing:.04em;
         text-transform:uppercase; }}
 .val {{ font-size:31px; font-weight:700; margin:.35rem 0 .45rem; letter-spacing:-.02em; }}
 .delta {{ font-size:11.5px; font-weight:700; }}
 .note {{ font-size:11.5px; color:var(--muted); margin-top:.2rem; }}
 h2 {{ font-size:13px; margin:0 0 .7rem; font-weight:700; letter-spacing:.02em;
       text-transform:uppercase; }}
 .ins {{ display:flex; gap:.7rem; padding:.75rem .85rem; border-radius:12px;
         background:var(--surface); border:1px solid var(--line); line-height:1.45; }}
 .dot {{ font-weight:800; font-size:12px; width:24px; height:24px; border-radius:8px;
         display:flex; align-items:center; justify-content:center; flex:0 0 24px; }}
 .head {{ font-size:13.5px; font-weight:600; color:var(--ink); }}
 .mean {{ font-size:12.5px; color:var(--ink2); margin-top:.25rem; line-height:1.5; }}
 .act {{ font-size:12.5px; color:var(--ink2); margin-top:.45rem; padding-top:.45rem;
         border-top:1px dashed var(--line); }}
 .act span {{ font-size:9.5px; font-weight:800; letter-spacing:.08em; margin-right:.45rem; }}
 table {{ border-collapse:collapse; width:100%; font-size:13px; }}
 th {{ text-align:right; padding:.55rem .75rem; border-bottom:2px solid var(--line);
       font-size:11px; color:var(--muted); font-weight:700; letter-spacing:.04em;
       text-transform:uppercase; white-space:nowrap; }}
 th:first-child,td:first-child {{ text-align:left; }}
 td {{ text-align:right; padding:.5rem .75rem; border-bottom:1px solid var(--hair);
       font-variant-numeric:tabular-nums; white-space:nowrap; }}
 tbody tr:nth-child(even) {{ background:var(--page); }}
 .scroll {{ overflow-x:auto; }}
 .muted {{ color:var(--muted); font-size:12px; }}
 footer {{ color:var(--muted); font-size:12px; margin-top:1.2rem; line-height:1.6; }}
 .noprint {{ position:fixed; top:14px; right:16px; z-index:99; display:flex;
             gap:.5rem; align-items:center; }}
 .btn {{ font-family:inherit; font-size:12.5px; font-weight:600; padding:.55rem .95rem;
         border-radius:10px; border:1px solid var(--ink); background:var(--ink);
         color:#fff; cursor:pointer; box-shadow:0 6px 18px -8px rgba(18,16,14,.5); }}
 .btn:hover {{ opacity:.88; }}
 .hint {{ font-size:11px; color:var(--muted); background:var(--surface);
          border:1px solid var(--line); padding:.35rem .55rem; border-radius:8px; }}
 @page {{ size:A4 landscape; margin:12mm 10mm; }}
 @media print {{
   .noprint {{ display:none !important; }}
   html,body {{ background:#fff; }}
   body {{ padding:0; -webkit-print-color-adjust:exact; print-color-adjust:exact; }}
   .wrap {{ max-width:none; }}
   .card,.tile,.ins {{ box-shadow:none; break-inside:avoid; page-break-inside:avoid; }}
   header,h2 {{ break-after:avoid; page-break-after:avoid; }}
   .kpis {{ grid-template-columns:repeat(4,1fr); }}
   .three,.two {{ grid-template-columns:repeat(2,1fr); }}
   .scroll {{ overflow:visible; }}
   table {{ font-size:11px; }}
   thead {{ display:table-header-group; }}
   tr {{ break-inside:avoid; page-break-inside:avoid; }}
 }}
</style></head><body><div class="wrap">
<div class="noprint"><span class="hint">Pick &ldquo;Save as PDF&rdquo; as the destination</span>
<button class="btn" onclick="window.print()">Download PDF</button></div>
<header><h1>{title}</h1><div class="cap">{caption}</div></header>
<div class="grid kpis">{tile_html}</div>
<div class="card" style="margin-top:.85rem"><h2>Insights</h2>
  <p class="muted" style="margin:-.35rem 0 .85rem">What moved, what it means, and what to
  do about it. Every figure is read from a query.</p>
  <div class="grid three" style="margin-top:0">{cards(findings)}</div>
</div>
<div class="grid">{"".join(chart_html)}</div>
<div class="grid">{"".join(table_html)}</div>
<footer><strong>How these numbers are built</strong>
<ul><li>Source: {SCHEMA}.{TABLE}</li>
<li>Measures: {", ".join(pretty(m) for m in MEASURES)}</li>
<li>Rates are averaged, never summed: {", ".join(pretty(r) for r in RATES) or "none"}</li>
</ul><strong>Not derivable from this dataset</strong><ul>{notes}</ul>
<p>Snapshot generated from a live Exasol query. Figures reflect the filters shown above.</p>
</footer></div>
<script>
 // Plotly sizes itself to the screen; without this the charts keep their screen
 // width on the printed page and spill off the sheet.
 window.addEventListener("beforeprint", function () {{
   if (!window.Plotly) return;
   document.querySelectorAll(".js-plotly-plot").forEach(function (node) {{
     window.Plotly.Plots.resize(node);
   }});
 }});
</script>
</body></html>\"\"\"


def build_insights(kpi, trend, comp, conc=None, cross=None):
    \"\"\"Findings, each carrying its own meaning and action.

    Every figure is read from a query result. The wording is templated, so a
    finding can never assert something the data does not contain.
    \"\"\"
    findings = []

    def add(text, tone, meaning=None, action=None):
        findings.append({"text": text, "tone": tone, "meaning": meaning, "action": action})

    headline = SERIES_MEASURES[0] if SERIES_MEASURES else (MEASURES[0] if MEASURES else "")
    moves = []
    for i, name in enumerate(MEASURES):
        cur, pri = _f(kpi.get(f"M{i}_CUR")), _f(kpi.get(f"M{i}_PRI"))
        if pri:
            moves.append((name, cur, pri, (cur / pri - 1) * 100))

    material = [m for m in moves if abs(m[3]) >= 3]
    for name, cur, pri, change in material:
        add(f"{pretty(name)} {'fell' if change < 0 else 'grew'} {abs(change):.1f}% to "
            f"{fmt(cur, name)}, against {fmt(pri, name)} a period earlier.",
            "bad" if change < 0 else "good",
            meaning=(f"A move of this size is outside normal period-to-period variation."),
            action=("Rebuild the forecast for the remaining period and brief the board on "
                    "the gap." if change < 0 else
                    "Confirm the gain is repeatable before building it into plan."))
    if not material and moves:
        add(f"Every headline measure is within 3% of the prior period — "
            f"{pretty(moves[0][0])} at {fmt(moves[0][1], moves[0][0])}.", "good",
            meaning="The business is holding its level; nothing here needs explaining "
                    "to a board.",
            action=None)

    named = {n: (c, p, d) for n, c, p, d in moves}
    # Substring matching alone picks a SUBSET of revenue -- "Loss-making sales"
    # contains "sales" -- and reports profit over it as the margin. Prefer an
    # exact name, then the largest candidate, which is the total the others
    # are carved out of.
    def pick(words):
        exact = [n for n in named if n.lower() in words]
        if exact:
            return exact[0]
        loose = [n for n in named if any(w in n.lower() for w in words)]
        return max(loose, key=lambda n: abs(named[n][0] or 0), default=None)

    profit = pick({"profit", "net profit", "gross profit", "margin", "net income"})
    revenue = pick({"sales", "revenue", "net sales", "net revenue", "gross revenue",
                    "turnover"})
    if profit == revenue:
        revenue = None
    if profit and revenue and named[revenue][0] and named[revenue][1]:
        margin_now = named[profit][0] / named[revenue][0] * 100
        margin_was = named[profit][1] / named[revenue][1] * 100
        shift = margin_now - margin_was
        add(f"Margin is {margin_now:.1f}% of revenue, {abs(shift):.2f}pp "
            f"{'down' if shift < 0 else 'up'} on the prior period.",
            "bad" if shift < -0.5 else ("good" if shift > 0.5 else "info"),
            meaning=("Revenue and profit are not moving together: price or cost is "
                     "eroding the margin, not only volume."
                     if shift < -0.5 else
                     "Revenue and profit are moving together, so the margin structure "
                     "is intact."),
            action=("Open a pricing and cost review on the largest groups before the "
                    "next forecast." if shift < -0.5 else None))

    if trend and len(trend) >= 6:
        values = [_f(r.get("M0")) for r in trend]
        recent, earlier = values[-3:], values[-6:-3]
        if sum(earlier):
            drift = (sum(recent) / 3) / (sum(earlier) / 3) - 1
            if abs(drift) >= 0.08:
                add(f"The last three periods run {abs(drift)*100:.0f}% "
                    f"{'below' if drift < 0 else 'above'} the three before them.",
                    "bad" if drift < 0 else "good",
                    meaning="Three periods in the same direction is a trend, not noise — "
                            "seasonality would have reverted by now.",
                    action=("Commission a root-cause review rather than waiting another "
                            "period." if drift < 0 else None))
        mean = sum(values) / len(values)
        if mean:
            spread = (max(values) - min(values)) / mean
            add(f"Period-to-period swing is {spread*100:.0f}% of the average across "
                f"{len(values)} periods.", "warn" if spread > 0.6 else "info",
                meaning=("Volatile enough that any single period is a poor guide — read "
                         "the moving average, not the last point."
                         if spread > 0.6 else
                         "Stable enough to plan against; a single period is representative."),
                action=("Set forecast ranges rather than point estimates."
                        if spread > 0.6 else None))

    if comp and len(comp) > 1:
        total = sum(_f(r.get("M0")) for r in comp) or 1
        top = comp[0]
        share = _f(top.get("M0")) / total * 100
        add(f"{pretty(top.get('SLICE'))} is the largest group at {share:.0f}% of "
            f"{pretty(headline)} ({fmt(top.get('M0'), headline)}).",
            "warn" if share >= 40 else "info",
            meaning=(f"At {share:.0f}% the aggregate is really this group's figure; "
                     f"movement elsewhere is invisible at this level."
                     if share >= 40 else
                     f"No group dominates — the total is a fair summary of all "
                     f"{len(comp)} of them."),
            action=(f"Report {pretty(top.get('SLICE'))} separately so the other groups "
                    f"stop being masked." if share >= 40 else None))

    if conc:
        values = sorted((_f(r.get("SHARE_VALUE")) for r in conc), reverse=True)
        total = sum(values) or 1
        top10 = sum(values[:10]) / total * 100
        add(f"The ten largest of {len(values):,} {ENTITY_NOUN} hold {top10:.1f}% of "
            f"{pretty(headline)}.",
            "bad" if top10 >= 40 else ("warn" if top10 >= 25 else "good"),
            meaning=(f"Losing one of the top ten would move the total by roughly "
                     f"{top10/10:.1f}% on its own."
                     if top10 >= 25 else
                     "No single account can move the total much — a genuinely broad base."),
            action=("Build a named retention plan for the top ten and quantify the "
                    "revenue at risk." if top10 >= 25 else None))

    if cross and len(cross) > 2:
        total = sum(_f(r.get("M0")) for r in cross) or 1
        best = max(cross, key=lambda r: _f(r.get("M0")))
        worst = min(cross, key=lambda r: _f(r.get("M0")))
        add(f"Strongest combination is {pretty(best.get('DIM_A'))} × "
            f"{pretty(best.get('DIM_B'))} at {_f(best.get('M0'))/total*100:.1f}% of the "
            f"total; weakest is {pretty(worst.get('DIM_A'))} × "
            f"{pretty(worst.get('DIM_B'))} at {_f(worst.get('M0'))/total*100:.2f}%.",
            "info",
            meaning="The spread between best and weakest cell shows where the mix is "
                    "carrying the number.",
            action=None)

    for cut in CUT_METRICS:
        add(f"{cut} was derived and then dropped.", "info",
            meaning="Flat across every dimension tested, so charting it would imply a "
                    "signal that is not there.", action=None)
    if REFUSED:
        add(f"{len(REFUSED)} question this persona asks cannot be answered from this "
            f"dataset.", "info",
            meaning="The required columns are absent — see the panel at the foot of "
                    "the page.",
            action="Point this persona at a dataset that carries them, or drop the "
                   "question from the pack.")
    return findings


def _run(server, metadata, sql):
    service = server.extensions.get("exasol_dashboard_service")
    if service is None:
        return None, "The Exasol service is not available in this context."
    profile = metadata["data_sources"]["primary"]["profile"]
    try:
        result = service.execute_profile_query(profile, sql, params={})
    except Exception as exc:
        return None, f"Query failed: {str(exc)[:200]}"
    rows = result.get("records") or result.get("rows") or []
    if rows and not isinstance(rows[0], dict):
        cols = [c if isinstance(c, str) else c.get("name") for c in result.get("columns", [])]
        rows = [dict(zip(cols, r)) for r in rows]
    return rows, None


QUESTION_NOISE = {"what", "whats", "what's", "is", "are", "the", "a", "an", "of", "in",
                  "for", "by", "to", "and", "or", "me", "show", "give", "list", "how",
                  "much", "many", "top", "best", "worst", "highest", "lowest", "total",
                  "per", "vs", "versus", "on", "at", "from", "with", "value", "values",
                  "please", "do", "we", "our", "there", "this", "that", "it", "was",
                  "average", "avg", "mean", "sum", "sums", "count", "over", "last",
                  "each", "all", "any", "which", "who", "where", "when", "have",
                  "overall", "total", "across", "per", "give", "tell", "get"}


def norm(text):
    \"\"\"Compare names on their words, not their punctuation.

    The panel *displays* "Weekly Sales" but the column is "Weekly_Sales", so raw
    matching could never match what the reader was told to type -- the refusal
    then asked for the exact words they had just used. Pad with spaces so a name
    matches on whole words: "Sales" must not match inside "wholesales".
    \"\"\"
    return " " + re.sub(r"[^a-z0-9]+", " ", str(text).lower()).strip() + " "


def match_values(text):
    \"\"\"Find dimension VALUES named in the question, longest literal first.

    "APAC" is a value of Market, not a column. Matching only column names lets a
    question's actual subject fall on the floor.
    \"\"\"
    # "top 5" is a row limit, not a value. Strip it before matching, or a
    # dimension whose values are numbers (Store 1-45, Pclass 1-3) turns
    # "top 5 store by sales" into a filter on store 5.
    text = re.sub(r"\\btop\\s+\\d{1,3}\\b", " ", text)
    hits, taken = [], set()
    for column, values in VALUES.items():
        best = None
        for value in values:
            low = value.lower()
            if low in taken or not low:
                continue
            # A bare number is only a value if the question also names its
            # column ("store 5"); otherwise it is a quantity, a year, a rank.
            if low.isdigit() and norm(column) not in norm(text):
                continue
            if re.search(r"\\b" + re.escape(low) + r"\\b", text) and (
                    best is None or len(low) > len(best[1])):
                best = (column, low, value)
        if best:
            hits.append((best[0], best[2]))
            taken.add(best[1])
    return hits


def leftover_terms(text, measure, dim, hits):
    \"\"\"Words the template could not account for -- the honest part of the answer.\"\"\"
    spoken = " ".join([measure or "", dim or ""] + [v for _, v in hits]).lower()
    spoken_words = set(re.findall(r"[a-z0-9]+", spoken))
    # Tokenise the question exactly as the spoken set was tokenised, or a hyphen
    # makes a term that WAS used look ignored ("sub-category" vs {"sub","category"}).
    words = re.findall(r"[a-z][a-z0-9]*", text)
    return [w for w in words
            if w not in QUESTION_NOISE and w not in spoken_words and len(w) > 2]


def _as_of(server, metadata):
    # Newest value of the fact table's time column, or None. Without it the
    # model answers "last year" with CURRENT_DATE, which on a historical
    # dataset falls outside the data and returns an empty table that reads
    # like a real answer. (Comment, not a docstring: this whole module is
    # itself a triple-quoted template.)
    if not HAS_TIME or not TIME_EXPR:
        return None
    # TIME_EXPR names the fact table by its own name as alias, so the source
    # has to be aliased the same way for the reference to resolve.
    source = f'"{SCHEMA}"."{TABLE}" "{TABLE}"'
    rows, error = _run(server, metadata,
                       f'SELECT MAX({TIME_EXPR}) AS D FROM {source}')
    if error or not rows:
        return None
    value = rows[0].get("D")
    return str(value)[:10] if value else None


def answer_question(server, metadata, text):
    t = (text or "").strip().lower()
    if not t:
        return None, None, "Ask something like: top 10 <dimension> by <measure>.", None
    proposed, model_error = _LLM.propose_sql(text, SCHEMA, TABLE, CATALOG, JOINS,
                                             as_of=_as_of(server, metadata))
    if proposed:
        safe, guard_error = _LLM.guard(proposed, SCHEMA, TABLE)
        if guard_error:
            return None, proposed, f"Generated SQL {guard_error}.", None
        rows, run_error = _run(server, metadata, safe)
        return (None, safe, run_error, None) if run_error else (rows, safe, None, None)

    # Longest name wins: "Category" is a substring of "Sub-Category", so first-match
    # order would answer a question about sub-categories with categories.
    question = norm(t)
    measure = max((m for m in MEASURES + RATES if norm(m) in question),
                  key=len, default=None)
    # CHAT_DIMS, not DIMS: this fallback emits a single-table query, so naming a
    # column that lives on a joined table would produce "object not found".
    dim = max((d for d in CHAT_DIMS if norm(d) in question), key=len, default=None)
    hits = match_values(t)
    # A value pins the row set; grouping by the column it came from would just
    # restate the filter, so drop that grouping.
    if dim and any(column == dim for column, _ in hits):
        dim = max((d for d in CHAT_DIMS if norm(d) in question
                   and not any(c == d for c, _ in hits)), key=len, default=None)
    n = 10
    match = re.search(r"top\\s+(\\d{1,3})", t)
    if match:
        n = min(int(match.group(1)), 200)

    src = f'"{SCHEMA}"."{TABLE}"'
    tick = chr(39)
    where = [f'"{column}" = {tick}{value.replace(tick, tick * 2)}{tick}'
             for column, value in hits]
    # The KPI cards report a trailing 12 months. Answering the same question
    # over all history puts two different numbers under one label on one page.
    windowed = bool(HAS_TIME and TIME_EXPR)
    if windowed:
        where.append(f"{TIME_EXPR} > ADD_MONTHS("
                     f"(SELECT MAX({TIME_EXPR}) FROM {src}), -12)")
    ignored = leftover_terms(t, measure, dim, hits)

    if not MEASURES:
        return None, None, "This dashboard has no measure to total.", None
    if measure is None:
        # Defaulting to MEASURES[0] answered "revenue by market" with loss-making
        # sales. Without a named measure there is no question to answer, and a
        # confident number against the wrong measure is worse than a refusal.
        wanted = ", ".join(pretty(m) for m in (MEASURES + RATES)[:6])
        unknown = (f" I could not match: {', '.join(sorted(set(ignored)))}."
                   if ignored else "")
        return None, None, (f"No measure named in the question, and no model is "
                            f"configured, so I will not guess one.{unknown} "
                            f"Name one of: {wanted}."), None

    agg = "AVG" if measure in RATES else "SUM"
    # A derived measure ("Slow shipments") names no column: quoting the name
    # produces "object not found". Always aggregate its expression.
    target = MEASURE_SQL.get(measure) or f'"{measure}"'
    if dim:
        clause = " AND ".join([f'"{dim}" IS NOT NULL'] + where)
        sql = (f'SELECT "{dim}", {agg}({target}) AS "{measure}" FROM {src} '
               f'WHERE {clause} GROUP BY "{dim}" ORDER BY 2 DESC LIMIT {n}')
    else:
        columns = ", ".join(f'"{column}"' for column, _ in hits)
        select = (columns + ", ") if columns else ""
        group = f' GROUP BY {columns}' if columns else ""
        clause = (" WHERE " + " AND ".join(where)) if where else ""
        sql = f'SELECT {select}{agg}({target}) AS "{measure}" FROM {src}{clause}{group}'

    safe, guard_error = _LLM.guard(sql, SCHEMA, TABLE)
    if guard_error:
        return None, sql, f"Template SQL {guard_error}.", None

    note = None
    if windowed:
        note = "Last 12 months, the same window as the cards above."
    if model_error:
        # Falling back silently made a broken model path look like a working one.
        note = ((note + " ") if note else "") + \
               f"The model path is configured but did not answer: {model_error}. " \
               "This answer came from template matching instead."
    if ignored:
        note = ((note + " ") if note else
                "No model is configured, so this was answered by template matching. ") + \
               (f"These words were not used: {', '.join(sorted(set(ignored)))}. "
                "Check the SQL before trusting the number.")
    rows, run_error = _run(server, metadata, safe)
    return (None, safe, run_error, None) if run_error else (rows, safe, None, note)


def workbook_bytes(sections):
    \"\"\"One sheet per section plus a provenance sheet, so numbers keep their context.\"\"\"
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    book = Workbook(); book.remove(book.active)
    header_fill = PatternFill("solid", fgColor="12100E")
    for sheet_name, rows in sections:
        sheet = book.create_sheet(sheet_name[:31])
        if not rows:
            sheet.append(["no rows"]); continue
        columns = list(rows[0].keys())
        sheet.append([pretty(c) for c in columns])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = header_fill
        for row in rows:
            sheet.append([row.get(c) for c in columns])
        for i, column in enumerate(columns, start=1):
            width = max(len(str(pretty(column))),
                        *(len(str(r.get(column, ""))) for r in rows[:200]))
            sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = \\
                min(max(width + 2, 12), 48)
    meta = book.create_sheet("About")
    for line in [["Dashboard", $title_r], ["Source schema", SCHEMA],
                 ["Fact table", TABLE], ["Persona", $persona_r],
                 ["Grain", $grain_label_r], [""],
                 ["Measures", ", ".join(pretty(m) for m in MEASURES)],
                 ["Rates (averaged, never summed)", ", ".join(pretty(r) for r in RATES)], [""],
                 ["Not derivable from this dataset"]] + [[n] for n in NEGATIVES + REFUSED]:
        meta.append(line)
    meta["A1"].font = Font(bold=True)
    meta.column_dimensions["A"].width = 34; meta.column_dimensions["B"].width = 60
    buffer = io.BytesIO(); book.save(buffer)
    return buffer.getvalue()


def pdf_bytes(title, caption, tiles, insight_groups, sections):
    \"\"\"A printable board pack: headline figures, what changed, then the tables.\"\"\"
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, PageBreak)

    ink = colors.HexColor("#12100e")
    muted = colors.HexColor("#8a837a")
    line = colors.HexColor("#e6e1d8")
    tone_colour = {"good": colors.HexColor("#0a7d4b"), "bad": colors.HexColor("#c0392b"),
                   "warn": colors.HexColor("#a8730a"), "info": muted}

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=20, textColor=ink,
                        alignment=0, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=muted,
                         spaceAfter=10)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=10, textColor=ink,
                        spaceBefore=10, spaceAfter=6)
    body_style = ParagraphStyle("body", parent=styles["Normal"], fontSize=9,
                                textColor=ink, leading=13)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm, title=title)
    flow = [Paragraph(title, h1), Paragraph(caption, sub)]

    if tiles:
        header = [Paragraph(f"<b>{t['label']}</b>", body_style) for t in tiles]
        values = [Paragraph(f"<font size=15><b>{t['value']}</b></font>", body_style)
                  for t in tiles]
        deltas = [Paragraph(t.get("delta") or "", body_style) for t in tiles]
        grid = Table([header, values, deltas], colWidths=[(268 / max(len(tiles), 1)) * mm] * len(tiles))
        grid.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LINEBELOW", (0, 2), (-1, 2), 0.5, line),
            ("TEXTCOLOR", (0, 0), (-1, 0), muted)]))
        flow += [grid, Spacer(1, 6)]

    observations, learnings, actions = insight_groups
    for heading, group in (("WHAT CHANGED", observations), ("WHAT IT MEANS", learnings),
                           ("ACTION REQUIRED", actions)):
        if not group:
            continue
        flow.append(Paragraph(heading, h2))
        for text, tone in group:
            flow.append(Paragraph(
                f'<font color="#{tone_colour.get(tone, muted).hexval()[2:]}">&#9632;</font> '
                f"{text}", body_style))
            flow.append(Spacer(1, 3))

    for name, rows in sections:
        if not rows:
            continue
        flow.append(PageBreak())
        flow.append(Paragraph(name.upper(), h2))
        columns = list(rows[0].keys())[:8]
        data = [[Paragraph(f"<b>{pretty(c)}</b>", body_style) for c in columns]]
        for row in rows[:26]:
            data.append([Paragraph(cell_value(c, row.get(c)), body_style)
                         for c in columns])
        # Explicit widths: without them a wide table silently squeezes to nothing.
        usable = 268 * mm
        first = usable * (0.28 if len(columns) > 2 else 0.5)
        rest = (usable - first) / max(len(columns) - 1, 1)
        table = Table(data, repeatRows=1,
                      colWidths=[first] + [rest] * (len(columns) - 1))
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#12100e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, line),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f5f1")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
        flow.append(table)
        if len(rows) > 26:
            flow.append(Paragraph(f"<i>showing 26 of {len(rows):,} rows — the Excel export "
                                  f"carries every row</i>", sub))

    flow.append(PageBreak())
    flow.append(Paragraph("HOW THESE NUMBERS ARE BUILT", h2))
    for label, value in (("Source schema", SCHEMA), ("Fact table", TABLE),
                         ("Measures", ", ".join(pretty(m) for m in MEASURES)),
                         ("Rates (averaged, never summed)",
                          ", ".join(pretty(r) for r in RATES) or "none")):
        flow.append(Paragraph(f"<b>{label}:</b> {value}", body_style))
        flow.append(Spacer(1, 3))
    if NEGATIVES + REFUSED:
        flow.append(Paragraph("NOT DERIVABLE FROM THIS DATASET", h2))
        for item in NEGATIVES + REFUSED:
            flow.append(Paragraph(f"&#8226; {item}", body_style))
            flow.append(Spacer(1, 2))
    doc.build(flow)
    return buffer.getvalue()


def create_dash_app(server, url_base_pathname, metadata):
    app = Dash(__name__, server=server, routes_pathname_prefix="/",
               requests_pathname_prefix=url_base_pathname.rstrip("/") + "/",
               external_stylesheets=[JAKARTA],
               title=metadata.get("title", $title_r), suppress_callback_exceptions=True)

    mount = url_base_pathname.rstrip("/")
    slug = mount.strip("/").replace("/", "_")
    snapshot_rule = "/__report/snapshot.html"
    snapshot_path = mount + snapshot_rule
    endpoint = "snapshot_" + slug
    if endpoint not in server.view_functions:
        def _serve_snapshot():
            from flask import Response
            page = _SNAPSHOT.get(mount)
            if not page:
                return Response("No snapshot yet — press Share on the dashboard first.",
                                mimetype="text/plain", status=404)
            return Response(page, mimetype="text/html; charset=utf-8")
        server.add_url_rule(snapshot_rule, endpoint, _serve_snapshot)

    button = {"fontSize": "12px", "fontWeight": 600, "padding": "0.45rem 0.85rem",
              "borderRadius": "9px", "border": "1px solid " + LINE, "background": SURFACE,
              "color": INK, "cursor": "pointer"}
    primary = {**button, "background": INK, "color": "#fff", "border": "1px solid " + INK}

    filter_controls = [html.Div([
        html.Label(pretty(name), style={"fontSize": "10.5px", "color": MUTED,
                                        "fontWeight": 700, "letterSpacing": "0.04em",
                                        "textTransform": "uppercase"}),
        dcc.Dropdown(id=f"filter-{i}", options=[], value=None,
                     placeholder="All " + pretty(name),
                     style={"minWidth": "185px", "fontSize": "13px"})])
        for i, name in enumerate(FILTERS)]

    header = html.Div([
        html.Div([
            html.Div($title_r, style={"fontSize": "24px", "fontWeight": 700, "color": INK,
                                      "letterSpacing": "-0.02em"}),
            html.Div(id="caption", style={"color": MUTED, "fontSize": "12.5px",
                                          "marginTop": "0.3rem"}),
        ]),
        html.Div(filter_controls + [
            html.Div([
                html.Button("Share", id="dl-share", n_clicks=0, style=primary),
                html.Div(id="share-link", style={"fontSize": "11px", "color": MUTED,
                         "maxWidth": "230px", "wordBreak": "break-all"}),
            ], style={"display": "flex", "gap": "0.4rem", "alignItems": "flex-end"}),
        ], style={"display": "flex", "gap": "0.7rem", "alignItems": "flex-end",
                  "flexWrap": "wrap"}),
    ], style={"display": "flex", "justifyContent": "space-between", "alignItems": "flex-end",
              "flexWrap": "wrap", "gap": "1.2rem", "padding": "1.4rem 1.6rem",
              "background": f"linear-gradient(180deg, {SURFACE} 0%, {PAGE} 100%)",
              "borderBottom": "1px solid " + LINE, "borderRadius": "16px 16px 0 0"})

    main = html.Div([
        html.Div(id="kpis", style={"display": "grid", "gap": "0.85rem",
                 "gridTemplateColumns": "repeat(auto-fit, minmax(196px, 1fr))"}),
        html.Div([
            html.H2("Insights", style=H2),
            html.Div("What moved, what it means, and what to do about it. Every figure "
                     "is read from a query — nothing here is generated prose.",
                     style={"fontSize": "12px", "color": MUTED, "margin": "0.35rem 0 0.85rem"}),
            html.Div(id="insights", style={"display": "grid", "gap": "0.55rem",
                     "gridTemplateColumns": "repeat(auto-fit, minmax(330px, 1fr))"}),
        ], style={**CARD, "marginTop": "0.85rem"}),
        html.Div(dcc.Graph(id="trend", config={"displayModeBar": False}),
                 style={**CARD, "marginTop": "0.85rem"}) if HAS_TIME else html.Div(),
        html.Div([
            html.Div(dcc.Graph(id="composition", config={"displayModeBar": False}),
                     style=CARD),
            html.Div(dcc.Graph(id="rates", config={"displayModeBar": False}), style=CARD),
        ], style={"display": "grid", "gap": "0.85rem", "marginTop": "0.85rem",
                  "gridTemplateColumns": "repeat(auto-fit, minmax(330px, 1fr))"}),
        html.Div([
            html.Div(dcc.Graph(id="concentration", config={"displayModeBar": False}),
                     style=CARD),
            html.Div([html.H2("Where the total comes from", style=H2),
                      html.Div(id="crosstab", style={"marginTop": "0.7rem"})], style=CARD),
        ], style={"display": "grid", "gap": "0.85rem", "marginTop": "0.85rem",
                  "gridTemplateColumns": "repeat(auto-fit, minmax(330px, 1fr))"}),
        html.Div([html.H2("Ranked detail", style=H2), html.Div(id="worklist",
                  style={"marginTop": "0.8rem"})],
                 style={**CARD, "marginTop": "0.85rem"}),
        html.Div([html.H2("Not derivable from this dataset", style=H2),
                  html.Ul([html.Li(x, style={"marginBottom": "0.2rem"})
                           for x in (NEGATIVES + REFUSED)] or
                          [html.Li("Nothing was refused for this persona.")],
                          style={"margin": "0.7rem 0 0", "paddingLeft": "1.1rem",
                                 "fontSize": "12px", "color": MUTED, "lineHeight": 1.6})],
                 style={**CARD, "marginTop": "0.85rem"}),
    ], style={"flex": "1 1 640px", "minWidth": 0})

    chat = html.Div([
        html.H2("Ask the data", style=H2),
        html.Div("Answers are generated as SQL, checked, then run read-only. "
                 "The query is always shown.",
                 style={"fontSize": "11.5px", "color": MUTED, "margin": "0.5rem 0 0.4rem",
                        "lineHeight": 1.5}),
        # Whoever installs this next needs to know the model is optional, that it
        # is theirs to supply, and where to put it -- from inside the app, not
        # from a README they may never open.
        html.Div([
            html.Span("No model key configured, so questions are matched on "
                      "keywords: plurals, synonyms and intent words such as "
                      "why or worst will not work. "),
            html.Span("To enable it, copy your Anthropic key and run ",
                      style={"opacity": 0.9}),
            html.Code(SETUP_KEY_CMD + " --clipboard", style={"fontSize": "10.5px",
                      "background": PAGE, "padding": "0.05rem 0.25rem",
                      "borderRadius": "4px", "border": "1px solid " + LINE}),
            html.Span(" — that route works anywhere, including an editor or "
                      "agent console. Run it with no flags in a real terminal "
                      "for a hidden prompt instead, or pass "),
            html.Code("--key-file PATH", style={"fontSize": "10.5px",
                      "background": PAGE, "padding": "0.05rem 0.25rem",
                      "borderRadius": "4px", "border": "1px solid " + LINE}),
            html.Span(". Don't paste the key into a chat: one that lands in a "
                      "transcript has to be rotated. No restart needed."),
        ], style={"fontSize": "11px", "color": WARN, "backgroundColor": WARN_BG,
                  "padding": "0.45rem 0.6rem", "borderRadius": "8px",
                  "margin": "0 0 0.75rem", "lineHeight": 1.5,
                  "display": "none" if MODEL_READY else "block"}),
        dcc.Input(id="q-input", type="text", debounce=True,
                  placeholder=f"e.g. {pretty(MEASURES[0]) if MEASURES else 'total'} by "
                              f"{pretty(DIMS[0]) if DIMS else 'group'}",
                  style={"width": "100%", "padding": "0.55rem 0.7rem", "fontSize": "13px",
                         "borderRadius": "9px", "border": "1px solid " + LINE,
                         "boxSizing": "border-box"}),
        html.Button("Run", id="q-run", n_clicks=0, style={**primary, "marginTop": "0.55rem"}),
        dcc.Loading(html.Div(id="q-out", style={"marginTop": "0.9rem"}),
                    type="dot", color=SERIES[0]),
    ], style={**CARD, "flex": "0 0 350px", "alignSelf": "flex-start",
              "position": "sticky", "top": "0.9rem"})

    app.layout = html.Div([
        dcc.Interval(id="boot", interval=350, n_intervals=0, max_intervals=1),
        dcc.Download(id="dl"),
        html.Div([header, html.Div([main, chat],
                 style={"display": "flex", "gap": "0.85rem", "alignItems": "flex-start",
                        "flexWrap": "wrap", "padding": "0.85rem"})],
                 style={"backgroundColor": PAGE, "border": "1px solid " + LINE,
                        "borderRadius": "16px", "overflow": "hidden"}),
    ], style={"fontFamily": FONT, "backgroundColor": PAGE, "padding": "1.4rem",
              "minHeight": "100vh", "boxSizing": "border-box"})

    if FILTERS:
        @app.callback([Output(f"filter-{i}", "options") for i in range(len(FILTERS))],
                      Input("boot", "n_intervals"))
        def _filters(_n):
            rows = load_rows(server, metadata, __file__, "queries/business/filters.sql")
            if has_error(rows):
                return [[] for _ in FILTERS]
            return [[{"label": str(r["DIM_VALUE"]), "value": str(r["DIM_VALUE"])}
                     for r in rows if r.get("DIM") == name] for name in FILTERS]

    def _load_all(values):
        params = {f"f{i}": (values[i] or "") if i < len(values) else ""
                  for i in range(len(FILTERS))}
        return (params,
                load_row(server, metadata, __file__, "queries/business/kpi.sql", params=params),
                load_rows(server, metadata, __file__, "queries/business/composition.sql",
                          params=params),
                load_rows(server, metadata, __file__, "queries/business/worklist.sql",
                          params=params),
                (load_rows(server, metadata, __file__, "queries/business/trend.sql",
                           params=params) if HAS_TIME else []),
                load_rows(server, metadata, __file__, "queries/business/concentration.sql",
                          params=params),
                load_rows(server, metadata, __file__, "queries/business/crosstab.sql",
                          params=params))

    @app.callback(
        Output("caption", "children"), Output("kpis", "children"),
        Output("insights", "children"), Output("composition", "figure"),
        Output("rates", "figure"), Output("concentration", "figure"),
        Output("crosstab", "children"), Output("worklist", "children"),
        *([Output("trend", "figure")] if HAS_TIME else []),
        Input("boot", "n_intervals"),
        *[Input(f"filter-{i}", "value") for i in range(len(FILTERS))])
    def _refresh(_n, *values):
        _params, kpi, comp, work, trend, conc, cross = _load_all(values)
        if has_error(kpi):
            panel = render_error_panel(kpi["_error"]); blank = empty_fig("")
            base = ["Query failed", panel, panel, blank, blank, blank, panel, panel]
            return tuple(base + ([blank] if HAS_TIME else []))

        headline_measure = (SERIES_MEASURES[0] if SERIES_MEASURES
                            else (MEASURES[0] if MEASURES else ""))
        scope = " · ".join([(values[i] or "All " + pretty(n))
                            for i, n in enumerate(FILTERS)]) or "All data"
        asof = kpi.get("ASOF")
        caption = ((f"As of {str(asof)[:10]}, last 12 months  ·  " if asof else "") +
                   f"{int(_f(kpi.get('ROWS_N'))):,} rows  ·  {scope}")

        tiles = []
        period = window_label(kpi.get("ASOF"))
        for i, name in enumerate(MEASURES):
            cur, pri = _f(kpi.get(f"M{i}_CUR")), _f(kpi.get(f"M{i}_PRI"))
            tiles.append(kpi_tile(pretty(name), fmt(cur, name), i,
                                  ((cur / pri - 1) * 100) if pri else None,
                                  f"prior {fmt(pri, name)}" if pri else None,
                                  period=period))
        for j, name in enumerate(RATES):
            tiles.append(kpi_tile(pretty(name), fmt_rate(kpi.get(f"R{j}_CUR"), name),
                                  len(MEASURES) + j, None, "averaged, never summed",
                                  period=period))

        findings = build_insights(kpi, trend, comp, conc, cross)
        insights = [insight_card(f) for f in findings]

        figures = make_figures(kpi, trend, comp, conc, headline_measure)
        cfig, rfig, kfig = (figures["composition"], figures["rates"],
                            figures["concentration"])

        if work and not has_error(work):
            named = [{"ENTITY": pretty(r.get("ENTITY")),
                      **{pretty(n): fmt(r.get(f"M{i}"), n)
                         for i, n in enumerate(SERIES_MEASURES)}}
                     for r in work[:25]]
            wl = data_table(named, ["ENTITY"] + [pretty(n) for n in SERIES_MEASURES],
                            raw=True)
        else:
            wl = html.Div("No rows.", style={"color": MUTED, "fontSize": "13px"})
        if cross and not has_error(cross):
            top_cells = sorted(cross, key=lambda r: -_f(r.get("M0")))[:24]
            total_cells = sum(_f(r.get("M0")) for r in cross) or 1
            col_a = pretty(DIMS[0]) if DIMS else "Group A"
            col_b = pretty(DIMS[1]) if len(DIMS) > 1 else "Group B"
            col_v, col_s = pretty(headline_measure), "Share"
            cross_rows = [{col_a: pretty(r.get("DIM_A")), col_b: pretty(r.get("DIM_B")),
                           col_v: fmt(_f(r.get("M0")), headline_measure),
                           col_s: f"{_f(r.get('M0'))/total_cells*100:.1f}%"}
                          for r in top_cells]
            crosstab = data_table(cross_rows, [col_a, col_b, col_v, col_s],
                                  aligns=["left", "left", "right", "right"], raw=True)
            crosstab = html.Div([crosstab], style={"maxHeight": "300px",
                                                   "overflowY": "auto"})
        else:
            crosstab = html.Div("No two-dimension breakdown available.",
                                style={"color": MUTED, "fontSize": "13px"})

        result = [caption, tiles, insights, cfig, rfig, kfig, crosstab, wl]
        if HAS_TIME:
            result.append(figures["trend"] or empty_fig("No time series."))
        return tuple(result)

    @app.callback(Output("dl", "data"), Output("share-link", "children"),
                  Input("dl-share", "n_clicks"),
                  *[State(f"filter-{i}", "value") for i in range(len(FILTERS))],
                  prevent_initial_call=True)
    def _download(_clicks, *values):
        _params, kpi, comp, work, trend, conc, cross = _load_all(values)
        rename = {f"M{i}": n for i, n in enumerate(SERIES_MEASURES)}
        rename.update({f"M{i}_CUR": f"{n} (current)" for i, n in enumerate(MEASURES)})
        rename.update({f"M{i}_PRI": f"{n} (prior)" for i, n in enumerate(MEASURES)})
        rename.update({f"R{i}_CUR": f"{n} (current)" for i, n in enumerate(RATES)})
        rename.update({f"R{i}_PRI": f"{n} (prior)" for i, n in enumerate(RATES)})
        rename.update({"ASOF": "As of", "ROWS_N": "Rows", "ENTITIES_N": "Distinct entities"})
        def clean(rows):
            return [] if (not rows or has_error(rows)) else [
                {rename.get(k, k): v for k, v in r.items()} for r in rows]
        sections = [("Summary", clean([kpi]) if kpi and not has_error(kpi) else []),
                    ("Trend", clean(trend)), ("Composition", clean(comp)),
                    ("Where it comes from", clean(cross)),
                    ("Ranked detail", clean(work))]
        scope = " · ".join([(values[i] or "All " + pretty(n))
                            for i, n in enumerate(FILTERS)]) or "All data"
        asof = kpi.get("ASOF") if kpi and not has_error(kpi) else None
        caption = ((f"As of {str(asof)[:10]}, last 12 months &middot; " if asof else "") +
                   f"{int(_f((kpi or {}).get('ROWS_N'))):,} rows &middot; {scope}")
        headline_measure = (SERIES_MEASURES[0] if SERIES_MEASURES
                            else (MEASURES[0] if MEASURES else ""))
        tiles = []
        for i, name in enumerate(MEASURES):
            cur, pri = _f(kpi.get(f"M{i}_CUR")), _f(kpi.get(f"M{i}_PRI"))
            change = ((cur / pri - 1) * 100) if pri else None
            tiles.append({"label": pretty(name), "value": fmt(cur, name),
                          "delta": (f"{'&#9650;' if change >= 0 else '&#9660;'} "
                                    f"{abs(change):.1f}% vs prior")
                                   if change is not None else "",
                          "good": (change or 0) >= 0,
                          "note": f"prior {fmt(pri, name)}" if pri else ""})
        findings = build_insights(kpi, trend, comp, conc, cross)
        figures = make_figures(kpi, trend, comp, conc, headline_measure)
        tables = [("Where the total comes from", clean(cross)[:40]),
                  ("Ranked detail", clean(work)[:40])]
        try:
            page = share_html($title_r, caption, tiles, findings, figures, tables)
        except Exception as exc:
            return (dict(content=f"Report generation failed: {exc}",
                         filename="share-error.txt"), "")
        _SNAPSHOT[mount] = page
        link = html.A("Open shareable link \u2197", href=snapshot_path,
                      target="_blank",
                      style={"color": SERIES[0], "fontWeight": 700,
                             "textDecoration": "none", "fontSize": "11.5px"})
        return dict(content=page, filename="$name-report.html"), link

    @app.callback(Output("q-out", "children"), Input("q-run", "n_clicks"),
                  Input("q-input", "value"), State("q-input", "value"),
                  prevent_initial_call=True)
    def _ask(_clicks, _typed, text):
        rows, sql, error, note = answer_question(server, metadata, text)
        blocks = []
        if sql:
            blocks.append(html.Details([
                html.Summary("SQL", style={"fontSize": "11px", "color": MUTED,
                                           "cursor": "pointer", "fontWeight": 700,
                                           "letterSpacing": "0.04em",
                                           "textTransform": "uppercase"}),
                html.Pre(sql, style={"fontSize": "11px", "background": PAGE,
                    "padding": "0.6rem", "borderRadius": "8px", "overflowX": "auto",
                    "border": "1px solid " + LINE, "whiteSpace": "pre-wrap",
                    "marginTop": "0.4rem", "color": INK_2})], open=False))
        if error:
            blocks.append(html.Div(error, style={"fontSize": "12px", "color": BAD,
                          "backgroundColor": BAD_BG, "padding": "0.5rem 0.6rem",
                          "borderRadius": "8px", "marginTop": "0.5rem"}))
        elif not rows:
            blocks.append(html.Div("The query ran and returned no rows.",
                          style={"fontSize": "12px", "color": MUTED,
                                 "marginTop": "0.5rem"}))
        else:
            if note:
                blocks.append(html.Div(note, style={"fontSize": "11.5px", "color": WARN,
                              "backgroundColor": WARN_BG, "padding": "0.5rem 0.6rem",
                              "borderRadius": "8px", "marginTop": "0.5rem"}))
            blocks.append(html.Div(data_table(rows[:25], list(rows[0].keys())),
                                   style={"marginTop": "0.6rem"}))
            blocks.append(html.Div(f"{len(rows)} row(s)", style={"fontSize": "11px",
                          "color": MUTED, "marginTop": "0.35rem"}))
        return blocks

    return app
""")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--card", required=True); ap.add_argument("--plan", required=True)
    ap.add_argument("--signal"); ap.add_argument("--name", required=True)
    ap.add_argument("--title", required=True); ap.add_argument("--out", required=True)
    args = ap.parse_args()

    card = json.load(open(args.card, encoding="utf-8")); card = card[0] if isinstance(card, list) else card
    plan = json.load(open(args.plan, encoding="utf-8"))
    signal = (json.load(open(args.signal, encoding="utf-8"))
              if args.signal else {"cut": []})
    m = Model(card, plan)

    queries = {"kpi.sql": sql_kpi(m), "composition.sql": sql_composition(m),
               "worklist.sql": sql_worklist(m), "filters.sql": sql_filters(m),
               "crosstab.sql": sql_crosstab(m), "concentration.sql": sql_concentration(m)}
    if m.time:
        queries["trend.sql"] = sql_trend(m)
    queries = {k: v for k, v in queries.items() if v.strip()}

    os.makedirs(os.path.join(args.out, "queries", "business"), exist_ok=True)
    for filename, text in queries.items():
        open(os.path.join(args.out, "queries", "business", filename), "w",
             encoding="utf-8").write(text)

    smoke = {f"queries/business/{k}": {p: "" for p in m.params()}
             for k in queries if "{f0!s}" in queries[k] or "{f1!s}" in queries[k]}
    open(os.path.join(args.out, "queries", "sql_smoke.json"), "w",
         encoding="utf-8").write(
        json.dumps(smoke, indent=2) + "\n")

    refused = [f"{item['decision']['statement']} — " +
               "; ".join(f"{why}" for _, why in item["reasons"])
               for item in plan.get("refused_decisions", [])]
    app_py = APP_TEMPLATE.substitute(
        title=args.title, title_r=repr(args.title),
        schema=m.schema, schema_r=repr(m.schema), table_r=repr(m.fact["name"]),
        name=args.name, persona=plan.get("persona", "unspecified"),
        persona_r=repr(plan.get("persona", "unspecified")),
        grain_label=m.grain_label, grain_label_r=repr(m.grain_label),
        measures=repr([c["name"] for c in m.measures]),
        series_measures=repr([c["name"] for c in m.series_measures()]),
        kinds=repr({c["name"]: measure_kind(c) for c in m.measures + m.rates}),
        measure_sql=repr({c["name"]: m.expr(c) for c in m.measures + m.rates}),
        rates=repr([c["name"] for c in m.rates]),
        dims=repr([c["name"] for c in m.dims]),
        chat_dims=repr([c["name"] for c in m.dims
                        if c.get("_table", m.fact["name"]) == m.fact["name"]]),
        values=repr({c["name"]: c["values"] for c in m.dims
                     if c.get("values")
                     and c.get("_table", m.fact["name"]) == m.fact["name"]}),
        filters=repr([c["name"] for c in m.filters]),
        has_time=repr(bool(m.time)),
        time_expr=repr(m.expr(m.time) if m.time else None),
        # Quoted and prefixed with the running interpreter: a bare path is not
        # executable on Windows, and "python3" can hit the Microsoft Store stub.
        setup_key_cmd=repr(setup_key_command()),
        entity_noun=repr(entity_noun(m)),
        refused=repr(refused),
        negatives=repr(plan.get("negative_coverage", [])),
        cut=repr([f"{v['aggregation']}({v['measure']})" for v in signal.get("cut", [])]),
        catalog=repr([{"table": m.fact["name"], "name": c["name"],
                       "role": c["semantic_role"], "additivity": c["additivity"]}
                      for c in m.fact["columns"]] +
                     [{"table": t, "name": c["name"], "role": c["semantic_role"],
                       "additivity": c["additivity"]}
                      for t in m.paths for c in m.tables[t]["columns"]
                      if c["semantic_role"] in ("categorical_dim", "state_flag",
                                                "entity_label", "monetary_amount")]),
        joins=repr([f'JOIN "{step["table"]}" ON "{step["table"]}"."{step["right_col"]}"'
                    f' = "{step["left"]}"."{step["left_col"]}"'
                    for path in m.paths.values() for step in path]))
    open(os.path.join(args.out, "app.py"), "w", encoding="utf-8").write(app_py)

    # Governed consumption outputs: each query becomes a downloadable dataset.
    outputs = [{"id": filename.replace(".sql", ""), "kind": "dataset",
                "classification": "internal",
                "title": filename.replace(".sql", "").title(),
                "source": {"type": "exasol_sql", "data_source": "primary",
                           "path": f"queries/business/{filename}"},
                "parameters": {"type": "object", "properties":
                               {p: {"type": "string"} for p in m.params()}}
                if f"queries/business/{filename}" in smoke else
                {"type": "object", "properties": {}},
                "formats": ["csv", "xlsx"]}
               for filename in queries]
    open(os.path.join(args.out, "dash-app.json"), "w",
         encoding="utf-8").write(json.dumps({
        "name": args.name, "title": args.title,
        "description": f"Derived for {plan.get('persona')} from {m.schema}.",
        "template": "exasol-analytics",
        "data_sources": {"primary": {"kind": "exasol", "profile": "starter-kit",
                                     "auth_mode": "local_direct"}},
        "consumption": {"outputs": outputs},
    }, indent=2) + "\n")
    os.makedirs(os.path.join(args.out, "assets"), exist_ok=True)
    open(os.path.join(args.out, "assets", "keep.txt"), "w",
         encoding="utf-8").write(
        "Snapshots written by the Share button land here and are served at "
        "<app>/assets/report.html\n")
    open(os.path.join(args.out, "llm_sql.py"), "w",
         encoding="utf-8").write(LLM_MODULE)
    open(os.path.join(args.out, "requirements.txt"), "w",
         encoding="utf-8").write(
        "dash>=4.0,<5.0\nplotly>=5.18\npyexasol>=2.2.2,<3.0\nanthropic>=1.0\n")
    print(f"workspace written to {args.out}: {len(queries)} queries + app.py + manifest")

    # Ask at the moment it matters. A new user reaches this line on their first
    # dashboard; a note in a README they may never open does not reach them.
    key_file = os.path.join(os.environ.get("EXAKIT_HOME",
                            os.path.join(os.path.expanduser("~"), ".exasol-starter-kit")),
                            "credentials", "anthropic_api_key")
    if not (os.environ.get("ANTHROPIC_API_KEY") or os.path.exists(key_file)):
        here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        print()
        print("  No model key is configured, so this dashboard's Ask-the-data panel")
        print("  will match keywords: plurals, synonyms and words like why or worst")
        print("  will not work. It is optional -- everything else runs without it.")
        print()
        setup = setup_key_command()
        print(f"  To enable semantic text-to-SQL, add your own Anthropic key.")
        print(f"  In a terminal, for a hidden prompt:")
        print(f"      {setup}")
        print(f"  Anywhere else (editor or agent console), copy the key first:")
        print(f"      {setup} --clipboard")
        print(f"      {setup} --key-file PATH")
        print("  Get a key at https://console.anthropic.com  (a question costs")
        print("  a fraction of a cent). No restart needed; it is read per question.")
        print()


if __name__ == "__main__":
    main()
